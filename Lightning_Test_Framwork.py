import os
import fnmatch
import sys
import commands
import csv
import datetime
import pymssql
from string import Template
from pprint import pprint
import jaydebeapi
import jpype
import time
from stat_validation import *
from fileLock import FileLock
from datamodelToInput import *
from sz_count_validation import *
import openpyxl
from openpyxl import Workbook
# from sql_dw_metadata_validation import *
from metadata_validation import *
# from denodo_count_validation import *
from logfile_validation import *
import time
import pyodbc

import sys
reload(sys)
sys.setdefaultencoding('utf8')


def getRowsAsDict(cursor) :
      import itertools
      desc = cursor.description
      column_names = [col[0] for col in desc]
      data = [dict(itertools.izip(column_names, row)) for row in cursor.fetchall()]
      return data

def getRowsAsDict(cursor) :
  import itertools
  desc = cursor.description
  column_names = [col[0] for col in desc]
  data = [dict(itertools.izip(column_names, ['' if v is None else v for v in row])) for row in cursor.fetchall()]
  return data


# ms_env = { '1' : {'host' :'10.66.24.155:1436','username' :'PMXliIDWDev','pass' : 'PMXliIDWDev' }, '2': {'host' :'10.66.24.155:1436','username' :'','pass' : '' } ,'3' : {'host' :'10.66.24.155:1436','username' :'','pass' : '' }}

# xlcatlin@vs-ssh.visualstudio.com:v3/xlcatlin/DEEP/test-automation-framework
# git clone xlcatlin@vs-ssh.visualstudio.com:v3/xlcatlin/DEEP/test-automation-framework
# git push -f xlcatlin@vs-ssh.visualstudio.com:v3/xlcatlin/DEEP/test-automation-framework test
ms_env = {
'xle-edw-build-dev' : { 'host' : '10.66.24.181:1436','username' : 'PMXLEDWDev', 'pass' : 'PMXLEDWDev' } ,\
'xle-mdm-hub-test' : { 'host' : '10.67.24.72:1436','username' : 'HDXLDEEPDev', 'pass' : 'SEcu$#re12#$%' } ,\
'xli-idw-build-qa' : { 'host' : '10.65.24.201:1436','username' : 'sqxledeep', 'pass' : '2sD9JxQM' } ,\
'xl-emiedb-analytics-dev' : { 'host' : '10.66.24.52:1436' ,'username' : 'PMXLDEEPDev', 'pass' : 'YcD4fFa5' } ,\
'xli-idw-stage-dev' : { 'host' : '10.66.24.45:1436','username' : 'PMXliIDWDev', 'pass' : 'PMXliIDWDev' } ,\
'xle-edw-stage-dev' : { 'host' : '10.66.24.155:1436','username' : 'PMXLEDWDev', 'pass' : 'PMXLEDWDev' } ,\
'EWRDB0408D' : { 'host' : '10.66.24.111:1436','username' : 'sqdeeptest', 'pass' : '!@#$sqdeeptest!@#$' },\
'xle-edw-stage-dev' : { 'host' : '10.66.24.155:1436','username' : 'PMXLEDWDev', 'pass' : 'PMXLEDWDev' } ,\
'xle-mdm-hub-qa' : { 'host' : '10.65.24.171:1436','username' : 'sqdeeptest', 'pass' : '$qdee4test' } ,\
'preprod' : { 'host' : 'xlc-azu-eus2-ppd-edsppd-dw-server.database.windows.net','username' : 'dwadmin', 'pass' : '22Ride@uber', 'database' : 'dw_xle_sz' },\
'LDGDEVLDWODS01' : {'host' : '10.1.0.72:1433','username' : 'staging_axa_acsUsr', 'pass' : 'as1226369$111'},
}


abc_ms_env = {
'deep-test-sql-server' : { 'host' : 'deep-test-sql-server','username' : 'xlcadmin@deep-test-sql-server', 'pass' : 'N0t@dm1n', 'database' : 'etl_dev_poc' },
'xlc-az-eus2-dev-eds-db-etladmin' : { 'host' : 'xlc-az-eus2-dev-eds-db-etladmin','username' : 'etladmin@xlc-az-eus2-dev-eds-db-etladmin', 'pass' : 'Xlc@admin', 'database' : 'etl_dev_poc' },
'xlc-az-eus2-ppd-edsppd-etlsqlserver' : { 'host' : 'xlc-az-eus2-ppd-edsppd-etlsqlserver','username' : 'etladmin@xlc-az-eus2-ppd-edsppd-etlsqlserver', 'pass' : 'Xlc@admin', 'database' : 'xlc-az-eus2-ppd-edsppd-db-etladmin' }
}

# user='xlcadmin@deep-test-sql-server'
# deep-test-sql-server.database.windows.net
# sqlcmd -S deep-test-sql-server.database.windows.net -U xlcadmin -P N0t@dm1n -Q "SELECT @@VERSION"
# sqlcmd -S deep-test-sql-server.database.windows.net -U etladmin -P Xlc@admin -Q "select convert(datetime,'2018-03-19 07:09:00.0',121)"

OutputFilePath = "../output/"
MetaOutputFilePath = "../output/metadata/"
# /opt/mssql-tools/bin/sqlcmd -S '10.66.24.155,1436' -U 'PMXLEDWDev' -P 'PMXLEDWDev' -Q 'SELECT top 10 name FROM master.sys.databases;' 
# check username and passwords in linux
# grep -rl 'pwd' $(grep -rl -m 100 '10.65.24.171' .)


# jdbc:hive2://zk2-eds3-c.v5fz4bobclbuxdi32zq5x3fice.cx.internal.cloudapp.net:2181,zk4-eds3-c.v5fz4bobclbuxdi32zq5x3fice.cx.internal.cloudapp.net:2181,zk6-eds3-c.v5fz4bobclbuxdi32zq5x3fice.cx.internal.cloudapp.net:2181/;serviceDiscoveryMode=zooKeeper;zooKeeperNamespace=hiveserver2

# mssql
def initialize_meta_mssql(servername):
  global ms_conn,ms_cursor
  ms_conn = pymssql.connect(server=ms_env[servername]['host'],user=ms_env[servername]['username'], password=ms_env[servername]['pass'])
  ms_cursor = ms_conn.cursor(as_dict=True)

#============================ Start Hive Configuration ==========================================================
def initialize_hive():
  global hive_conn,hive_curs,Environment
  if not jpype.isJVMStarted() :
    jHome = jpype.getDefaultJVMPath()

    class_path = str.join(":", [os.path.abspath("../driver_jars/"+file) for file in os.listdir("../driver_jars")])
    jpype.startJVM(jHome, '-ea',  '-Djava.class.path='+class_path)

  #configuring log4j,otherwise will throw warning saying it isn't configured
  log4j = jpype.JPackage('org.apache.log4j')

  log4j.BasicConfigurator.configure()
  #setting the rootlogger to display only error messages
  log4j_logger = log4j.Logger.getRootLogger()
  log4j_logger.setLevel(log4j.Level.ERROR)

  driverName = "org.apache.hive.jdbc.HiveDriver";

  #project and environement specific variables

  # server for hive https://acen03-deep-hdi-hive02-d.azurehdinsight.net/
  # url = 'jdbc:hive2://10.186.224.51:10001/default;transportMode=http;httpPath=/hive2'

  # new server for hive https://acen11-deep-hdi-hive04-d.azurehdinsight.net/
  # url = 'jdbc:hive2://10.186.224.169:10001/default;transportMode=http;httpPath=/hive2'
  # user = 'sshuser'
  # passd ='D33ppassword123!'
  
  Environment = os.environ['env'] if os.environ['env'] else os.environ['ENV']
  # Environment = 'test'
  # if int(Environment) == 1:
  if Environment.strip().lower() == "random":
    # exclusive qa server for hive https://acen11-deep-hdi-hive04-d.azurehdinsight.net/
    url = 'jdbc:hive2://acen13-hdi-deep-hivellap-d.azurehdinsight.net:443/default;transportMode=http;ssl=true;httpPath=/hive2'
  # elif int(Environment) == 2:
  elif Environment.strip().lower() == "dev":
    # new server for strategic dev
    url = 'jdbc:hive2://10.185.195.54:10001/default;encrypt=true;httpPath=/hive2'
    user = 'qausr'
    passd ='aA1234567890'
  # elif int(Environment) == 3:
  elif Environment.strip().lower() == "test":
    # new server for strategic test
    # url = 'jdbc:hive2://10.185.196.218:10001/default;transportMode=http;httpPath=/hive2' 
    # updated stet server details
    url = 'jdbc:hive2://10.185.203.38:10001/default;encrypt=true;transportMode=http;httpPath=/hive2'
    user = 'sshuser'
    passd ='aA1234567890'
  # elif int(Environment) == 4:
  elif Environment.strip().lower() == "uat":
    # new server for strategic test
    url = 'jdbc:hive2://10.185.135.57:10001/default;transportMode=http;httpPath=/hive2'
    user = 'sshuser'
    passd ='aA1234567890'
  # elif int(Environment) == 5:
  elif Environment.strip().lower() == "prod":
    # new server for prod
    url = 'jdbc:hive2://10.185.138.53:10001/default;transportMode=http;httpPath=/hive2'
    user = 'qausr'
    passd ='aA1234567890'

  # user = 'hiveusr'
  # passd = 'hiveusr'

  # user = 'qausr'
  # passd = 'qausr'
  # print(url+" user: "+user+" pass: "+passd)

  hive_conn = jaydebeapi.connect(driverName,url,[user, passd])
  hive_curs = hive_conn.cursor()
  print("connected to hive server")

def close_hive() :
  global hive_conn,hive_curs
  hive_curs.close()
  hive_conn.close()
def get_mssql_structure(database,schema_name,tablename) :
    global ms_conn,ms_cursor
    try :
        ms_cursor.execute(sql_string.substitute({'databasename':database,'tablename':tablename ,'schemaname': schema_name}))
    except Exception as e :
        print(sql_string.substitute({'databasename':database,'tablename':tablename ,'schemaname': schema_name}))
        print(str(e))

    return ms_cursor
#============================ End Hive Configuration ==========================================================
#============================Print Summary Code redundancy removal ============================================
def printsummary(validation,count,passed,failed,ParasFilePath,countOutputFilename,ExecuStartTime) :
  ostring ="\n================================ "+validation+" Validation Summary ===============================================\n"\
   +" \nTotal number of Tables Validated                    : " + str(count)\
   +" \nNumber of Tables that passed "+validation+" validation : " + str(passed)\
   +" \nNumber of Tables that failed "+validation+" validation : " + str(failed)\
   +"\n"\
   +" \nInput parameter file        : " + ParasFilePath\
   +" \nOutput File has been stored : " + countOutputFilename\
   +" \nEnvironment selected        : " + Environment\
   +" \nExecution started Time      : " + ExecuStartTime\
   +" \nExecution End Time          : " + str(datetime.datetime.now())\
   +"\n================================================================================================\n"
  return ostring
#============================ Excel to Dict ===================================================================

def XLSXDictReader(f):
    book = openpyxl.reader.excel.load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))
    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)
    return (headers.values(),(dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1)))

def XLSXDictReader1(f):
  from collections import OrderedDict
  book = openpyxl.reader.excel.load_workbook(f)
  sheet = book.active
  rows = sheet.max_row
  cols = sheet.max_column
  headers = OrderedDict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols+1))
  def item(i, j): 
    return (sheet.cell(row=1, column=j).value,sheet.cell(row=i, column=j).value)
  
  return (headers.values(),[OrderedDict([item(i, j) for j in range(1, cols+1)]) for i in range(2,rows+1) if sheet.cell(row=i, column=1).value is not None] )

#============================ Excel to Dict end ===================================================================


#============================ Start neteeza Configuration ==========================================================
def initialize_neteeza(Environment=1):
  Environment = os.environ['env'] if os.environ['env'] else os.environ['ENV']
  if Environment.strip().lower() == "dev":
    dsn_database = "DEV5_ODS_DEEP" #DEV5_ODS_DEEP          
    dsn_hostname = "10.0.6.53"  
    dsn_port = "5480"                
    dsn_uid = "ee_user_pov"        
    dsn_pwd = "XLCatl1n"

  elif Environment.strip().lower() == "test":
    dsn_database = "STEST_FRAME_ODS_DEEP"       
    dsn_hostname = "10.0.6.53"  
    dsn_port = "5480"                
    dsn_uid = "STEST_FRAME"#  "SQLDEEPDEV1"      
    dsn_pwd = 'T$3st@fr@m' 

  elif Environment.strip().lower() == "uat":
    dsn_database = "PREPROD_FRAME_ODS_DEEP"       
    dsn_hostname = "10.0.6.53"  
    dsn_port = "5480"                
    dsn_uid = "PREPROD_FRAME"#  "SQLDEEPDEV1"      
    dsn_pwd = 'Pr#39r0d_fr@m3'  



  global neteeza_conn,neteeza_curs
  if not jpype.isJVMStarted() :
    jHome = jpype.getDefaultJVMPath()
    class_path = str.join(":", [os.path.abspath("../driver_jars/"+file) for file in os.listdir("../driver_jars")])
    jpype.startJVM(jHome, '-ea',  '-Djava.class.path='+class_path)

    #configuring log4j,otherwise will throw warning saying it isn't configured
    log4j = jpype.JPackage('org.apache.log4j')

    log4j.BasicConfigurator.configure()
    #setting the rootlogger to display only error messages
    log4j_logger = log4j.Logger.getRootLogger()
    log4j_logger.setLevel(log4j.Level.ERROR)

  driverName = "org.netezza.Driver";

  connection_string='jdbc:netezza://'+dsn_hostname+':'+dsn_port+'/'+dsn_database
  url = '{0}:user={1};password={2}'.format(connection_string, dsn_uid, dsn_pwd)
  # print("URL: " + url)
  # print("Connection String: " + connection_string)

  neteeza_conn = jaydebeapi.connect(driverName,connection_string,[dsn_uid, dsn_pwd])
  neteeza_curs = neteeza_conn.cursor()
  # print("connected to neteeza server")

def closeneteeza() :
  global neteeza_conn,neteeza_curs
  neteeza_conn.close()
  neteeza_curs.close()

#============================ End neteeza Configuration ==========================================================




#============================ start abc MSSQL Configuration ==========================================================
def initialize_mssql(servername):
  global ms_conn,ms_cursor
  ms_conn = pymssql.connect(server=ms_env[servername]['host'],user=ms_env[servername]['username'], password=ms_env[servername]['pass'])
  # database=ms_env[servername]['host'] optional parameter database to connect to 
  ms_cursor = ms_conn.cursor(as_dict=True)
  
def initialize_abcmssql(servername=None):
  global abc_ms_conn,abc_ms_cursor
  server = 'xlc-azu-eus2-ppd-edsppd-etlsqlserver.database.windows.net'
  database = 'xlc-azu-eus2-ppd-edsppd-db-etladmin'
  username = 'etladmin'
  password = 'Xlc@admin'
  driver= '{ODBC Driver 17 for SQL Server}'
  abc_ms_conn = pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)    
  abc_ms_cursor = abc_ms_conn.cursor()
  # abc_ms_conn = pymssql.connect(server=abc_ms_env[servername]['host'],user=abc_ms_env[servername]['username'], password=abc_ms_env[servername]['pass'],database=abc_ms_env[servername]['database'])
  # abc_ms_cursor = abc_ms_conn.cursor(as_dict=True)


def close_mssql():
  global ms_conn,ms_cursor
  ms_cursor.close()
  ms_conn.close()  

def close_abcmssql():
  global abc_ms_conn,abc_ms_cursor
  abc_ms_conn.close()
  abc_ms_cursor.close()  
  
#============================ End MSSQL Configuration ==========================================================

ExecuStartTime = ""
RecordCount = 0
RecordCount_Insert = 0
RecordCount_Delete = 0

def checkFileValid(recordlist) :
  if any( e['LoadDate'].strip() =='' or e['LoadDate'] is None  for e in recordlist) :
    pass
    # print("The loaddate Column is a required columns,values are missing") 
    # sys.exit()

  if any( e['LZ_BatchId'].strip() =='' or e['LZ_BatchId'] is None  for e in recordlist) :
    pass
    # print("The LZ_BatchId Column is a required columns,values are missing") 
    # sys.exit()   

  if any( e['OZ_BatchId'].strip() =='' or e['OZ_BatchId'] is None  for e in recordlist) :
    pass
    # print("The OZ_BatchId Column is a required columns,values are missing") 
    # sys.exit()    
  return



def formatABCOutput(cursor) :
  # print(resultset)

  lz_abc_string = ""
  oz_abc_string = ""
  resultset = getRowsAsDict(cursor)
  if resultset is None :
    return ""

  for row in resultset :
    if row['batch_set_category'] == 'landing' :
      lz_abc_string = str(row['batch_key'])+","+str(row['measure_src_value'])+","+str(row['measure_trgt_value'])+","+str(row['batch_sts_descr'])+","+str(row['balanced_ind'])
    elif row['batch_set_category'] == 'operational' :
      oz_abc_string = str(row['batch_key'])+","+str(row['measure_src_value'])+","+str(row['measure_trgt_value'])+","+str(row['batch_sts_descr'])+","+str(row['balanced_ind'])
  return lz_abc_string +"," + oz_abc_string

# batch ,record in src ,records in tgt,indicator
# measure_src_value
# measure_trgt_value
# batch_set_sts_descr
# balanced_ind
# measure_type='Count'

# batch_set_key   batch_key       batch_set_src_sys_nme   batch_set_category      batch_set_sts_descr     batch_src_db_nme        batch_trgt_db_nme       batch_nme       batch_sts_descr measure_type    measure_src_value       measure_trgt_value   balanced_ind    RN
# 210     314     conformance_mdm landing success xl_mdm_hub_ITC  xl_mdm_hub_mdm_lz       v_EDW_ContractType      success Count   46.00000000     46.00000000     Y       1
# 212     317     conformance_mdm operational     success xl_mdm_hub_mdm_lz       xl_mdm_hub_mdm_oz       v_EDW_ContractType      success Count   46.00000000     46.00000000     Y       1




def getABC_Status(batchid) :
  sql_query = 'SELECT B_SET.batch_set_key as batch_set_key,BAT.batch_key as batch_key,batch_set_src_sys_nme,batch_set_category,batch_set_sts_descr,batch_src_db_nme,batch_trgt_db_nme,batch_nme,batch_sts_descr,measure_type,measure_src_value,measure_trgt_value,balanced_ind FROM batch.t_batch_set B_SET JOIN batch.t_batch BAT ON B_SET.batch_set_key=BAT.batch_set_key JOIN audit.t_balance BAL ON BAT.batch_key=BAL.batch_key WHERE BAL.measure_type=\'Count\' and BAT.batch_key ={}'.format(batchid)
  try:
    start_time = time.time()
    abc_ms_cursor.execute(sql_query)
    print("Execution Time : {} seconds.".format(time.time() - start_time))
  except pyodbc.OperationalError as ex :
    sqlstate = ex.args[0]
    if sqlstate == '08S01' :      
      initialize_abcmssql()
      abc_ms_cursor.execute(sql_query)
      formatABCOutput(abc_ms_cursor)
  except Exception as ex :
    print(str(ex))
  
  # row = abc_ms_cursor.fetchall()

  return formatABCOutput(abc_ms_cursor)


def getABC_StatusTableName(tablename) :
  sql_query = 'SELECT * FROM(SELECT A.batch_set_key as batch_set_key,A.batch_key as batch_key,batch_set_src_sys_nme,batch_set_category,batch_set_sts_descr,batch_src_db_nme,batch_trgt_db_nme,batch_nme,batch_sts_descr,measure_type,measure_src_value,measure_trgt_value,balanced_ind,ROW_NUMBER()OVER(PARTITION BY batch_nme,batch_src_db_nme ORDER BY batch_strt_tm DESC)AS RN FROM batch.t_batch A JOIN(SELECT batch_set_key,batch_set_category,batch_set_src_sys_nme,batch_set_sts_descr FROM batch.t_batch_set WHERE batch_set_category IN(\'landing\',\'operational\'))B ON A.batch_set_key=B.batch_set_key JOIN audit.t_balance BAL ON A.batch_key=BAL.batch_key WHERE BAL.measure_type=\'Count\' and A.batch_nme =\'{}\')C WHERE C.RN=1'.format(tablename)
  # print(sql_query)
  try:
    # initialize start time
    start_time = time.time()
    abc_ms_cursor.execute(sql_query)
    print("ABC Execution Time : {} seconds.".format(time.time() - start_time))
  except pyodbc.OperationalError as ex :
    sqlstate = ex.args[0]
    if sqlstate == '08S01' :      
      initialize_abcmssql()
      abc_ms_cursor.execute(sql_query)
      formatABCOutput(abc_ms_cursor)
  except Exception as ex :
    print(str(ex))
  
  # row = abc_ms_cursor.fetchall()

  return formatABCOutput(abc_ms_cursor)


def getsqlrowcount(database,schemaname,tablename,CDC_Columns=None,CDC_Dates=None,checkCDC=False) :
  if checkCDC :
    try :
      if CDC_Dates :
        where_clause = " or ".join([ dates[0] +">'"+dates[1]+"'" for dates in zip(CDC_Columns.split(','),CDC_Dates.split(',')) ])
        # print(" and ".join([ dates[0] +">='"+dates[1]+"'" for dates in zip(CDC_Columns.split(','),CDC_Dates.split(',')) ]))
        # print("select count(*) as count from {}.{}.{} where {};".format(database,schemaname,tablename,where_clause))
        ms_cursor.execute("select count(*) as count from {}.{}.{} where {};".format(database,schemaname,tablename,where_clause))
        row = ms_cursor.fetchone()
        return row['count']         
      else :
        print("please run cdc update or update the cdc date columns manually for table : "+database+"."+schemaname+"."+tablename)
        # ms_cursor.execute("select count(*) as count from {}.{}.{} where dw_create_dt >='{}' and dw_last_update_dt>='{}';".format(database,schemaname,tablename,createtime,updatetime))
        # # print("select count(*) as count from {}.{}.{} where dw_create_dt >='{}' and dw_last_update_dt>='{}';".format(database,schemaname,tablename,createtime,updatetime))
        # row = ms_cursor.fetchone()
        return -1
    except pymssql.ProgrammingError as e : 
      if "Invalid column name" in str(e) :
        # we get the list of columns
        ms_cursor.execute("select top 1 * from {}.{}.{} ;".format(database,schemaname,tablename))
        desc = ms_cursor.description
        column_names = [col[0] for col in desc]
        column_names = map(lambda x : x.lower(),column_names)

        print("Column Names available for "+database+","+schemaname+","+tablename+": \n "+",".join(column_names))
        # we get the list of columns available from our intrested list
        # time_cols_avl = set(column_names).intersection(set(['dw_create_dt','dw_last_update_dt','datecreated','datemodified']))
        # time_cols_avl = set(column_names).intersection(set(['activationdatekey','dateadded','datecreated','datemodified','deactivationdatekey','dw_create_dt','dw_insert_dt','dw_last_update_dt','remove_event_dt','remove_event_tm']))

        # # building the select string 
        # sql_string = "select count(*) as count"
        # sql_string +=" from {}.{}.{} where ".format(database,schemaname,tablename)

        # if 'dw_create_dt' in time_cols_avl and 'dw_last_update_dt' in time_cols_avl :
        #   sql_string +=" dw_create_dt >= convert(datetime,'"+createtime+"',121) and dw_last_update_dt >= convert(datetime,'"+updatetime+"',121)"

        # elif'datecreated' in time_cols_avl and 'datemodified' in time_cols_avl:
        #   sql_string +=" datecreated >= convert(datetime,'"+createtime+"',121) and datecreated >= convert(datetime,'"+updatetime+"',121)"

        # else :
        #   print('Expected columns missing in MSSQL',column_names,time_cols_avl)
        #   return 0
        #   # raise ValueError('Expected columns missing in MSSQL',column_names,time_cols_avl)

        # # print(sql_string)
        # ms_cursor.execute(sql_string)
        # row = ms_cursor.fetchone()

        return -1
      # if this is a different error other than a invalid column then raise the error
      else :
        print("expection in getting Count from MSSQl for {}.{}.{} ".format(database,schemaname,tablename)+"\n"+str(e))
        return -1       
  else :
    try :
      # initialize start time
      start_time = time.time()
      ms_cursor.execute("select count(*) as count from {}.{}.{}".format(database,schemaname,tablename))
      # print("select count(*) as count from {}.{}.{}".format(database,schemaname,tablename))
      row = ms_cursor.fetchone()
      print("SQL Execution Time : {} seconds.".format(time.time() - start_time))
      return row['count']
    except Exception  as e :
      print("expection in getting Count from MSSQl for {}.{}.{} ".format(database,schemaname,tablename)+"\n"+str(e))
      return -1


def getneteezarowcount(database,schemaname,tablename,CDC_Columns=None,CDC_Dates=None,checkCDC=False) :
  if checkCDC :
    try :
      if CDC_Dates :
        where_clause = " or ".join([ dates[0] +">'"+dates[1]+"'" for dates in zip(CDC_Columns.split(','),CDC_Dates.split(',')) ])
        # print(" and ".join([ dates[0] +">='"+dates[1]+"'" for dates in zip(CDC_Columns.split(','),CDC_Dates.split(',')) ]))
        # print("select count(*) as count from {}.{}.{} where {};".format(database,schemaname,tablename,where_clause))
        neteeza_curs.execute("select count(*) as count from {}.{}.{} where {};".format(database,schemaname,tablename,where_clause))
        row = neteeza_curs.fetchone()
        return row[0]
      else :
        print("please run cdc update or update the cdc date columns manually for table : "+database+"."+schemaname+"."+tablename)
        # neteeza_curs.execute("select count(*) as count from {}.{}.{} where dw_create_dt >='{}' and dw_last_update_dt>='{}';".format(database,schemaname,tablename,createtime,updatetime))
        # # print("select count(*) as count from {}.{}.{} where dw_create_dt >='{}' and dw_last_update_dt>='{}';".format(database,schemaname,tablename,createtime,updatetime))
        # row = neteeza_curs.fetchone()
        return -1
    except Exception as e : 
      if "ERROR:  Attribute " in str(e) :
        # we get the list of columns
        neteeza_curs.execute("select * from {}.{}.{} limit 1;".format(database,schemaname,tablename))
        desc = neteeza_curs.description
        column_names = [col[0] for col in desc]
        column_names = map(lambda x : x.lower(),column_names)
        print("Column Names available for "+database+","+schemaname+","+tablename+": \n "+",".join(column_names))
        return -1
      # if this is a different error other than a invalid column then raise the error
      else :
        print("excepction in getting Count from Neteeza for {}.{}.{} ".format(database,schemaname,tablename)+"\n"+str(e))
        return -1       
  else :
    try :
      neteeza_curs.execute("select count(*) as count from {}.{}.{}".format(database,schemaname,tablename))
      # print("select count(*) as count from {}.{}.{}".format(database,schemaname,tablename))
      row = neteeza_curs.fetchone()
      # print(row)
      return row[0]
    except Exception  as e :
      print("expection in getting Count from Neteeza for {}.{}.{} ".format(database,schemaname,tablename)+"\n"+str(e))
      return -1

def gethiveozcount(schemaname,tablename) :
    try:
      hive_curs.execute("select count(*) as count from {}.{}".format(schemaname,tablename))
      # print("select count(*) as count from {}.{}".format(schemaname,tablename))
      row = hive_curs.fetchone()
      return row[0]
    except Exception as e:
      print("Error : "+schemaname+" "+tablename)
      print(str(e))

def gethiveszcount(schemaname,tablename) :
  return gethiveozcount(schemaname,tablename)

def gethiveszsqlcount(sql) :
  try:
    hive_curs.execute(sql)
    # print("select count(*) as count from {}.{}".format(schemaname,tablename))
    row = hive_curs.fetchone()
    return row[0]
  except Exception as e:
    print(str(e))


def gethivelzcount(schemaname,tablename,loaddate=None) :
  try:
    if loaddate is None or loaddate == "None":
      # hive_curs.execute("select max(loaddate) as loaddate from {}.{}".format(schemaname,tablename))
      # # print("select max(loaddate) as loaddate from {}.{}".format(schemaname,tablename))
      # row = hive_curs.fetchone()    
      # maxloaddate = row[0]
      # # print("hive: "+schemaname+","+tablename+" loaddate: "+str(maxloaddate))
      # hive_curs.execute("select max(loaddate) as loaddate,count(*) from {0}.{1} ".format(schemaname,tablename))
      # print("select max(loaddate) as loaddate,count(*) from {0}.{1} ".format(schemaname,tablename))
      start_time = time.time()
      hive_curs.execute("select loaddate,count(*) as count from {0}.{1} group by loaddate order by loaddate desc limit 1".format(schemaname,tablename))
      print("Loaddate Execution Time : {} seconds.".format(time.time() - start_time))
      # print("select loaddate,count(*) as count from {0}.{1} group by loaddate order by loaddate desc limit 1".format(schemaname,tablename))
      row = hive_curs.fetchall()
      return(row[0][0],row[0][1])

    else :
      start_time = time.time()
      hive_curs.execute("select count(*) as count from {}.{} where loaddate={}".format(schemaname,tablename,loaddate))
      print("Hivesql Execution Time : {} seconds.".format(time.time() - start_time))
      # print("select count(*) as count from {}.{} where loaddate={}".format(schemaname,tablename,loaddate))
      row = hive_curs.fetchone()
      return row[0]
  except Exception as e:
    print("Error : "+schemaname+" "+tablename)
    print(str(e))
    return -1

def getcdchive(schemaname,tablename,loaddate=None,column_names=None) :
  SecondMax_loaddate = 0
  if loaddate is None or loaddate == "None":
    try :
      hive_curs.execute("select distinct loaddate from {0}.{1} order by loaddate desc limit 2".format(schemaname,tablename))
      loaddate_rows = hive_curs.fetchall()
      if len(loaddate_rows) == 1 :
        print(loaddate_rows[0][0])
        SecondMax_loaddate = loaddate_rows[0][0]
        # if only one loaddate is present then dont take cdcdates
        print("Only One loaddate found :"+schemaname+" "+tablename+"" +loaddate)
        return {'' :'',"loaddate":loaddate_rows[0][0]}
      elif len(loaddate_rows) == 2 :
        SecondMax_loaddate = loaddate_rows[1][0]
      else :
        return {"error no load date found" :"error no load date found","loaddate":"no loaddate found"}
        print("select distinct loaddate from {0}.{1} order by loaddate desc limit 2".format(schemaname,tablename))
    except Exception as e:
      print("Error : {} {} ".format(schemaname,tablename))
      print(str(e))
      return {"sql error" :"sql error","loaddate":"error"}

  else :
    try :
      hive_curs.execute("select distinct loaddate from {0}.{1} where loaddate <={2} order by loaddate limit 2".format(schemaname,tablename,loaddate))
      loaddate_rows = hive_curs.fetchall()

      if len(loaddate_rows) == 1 :
        SecondMax_loaddate = loaddate_rows[0][0]
        # if only one loaddate is present then dont take cdcdates
        return {'only one load present' :'only one load present',"loaddate":loaddate_rows[0][0]}
      elif len(loaddate_rows) == 2 :
        SecondMax_loaddate = loaddate_rows[1][0]
      else :
        return {"error no load date found" :"error no load date found","loaddate":"no loaddate found"}
    except Exception as e:
      print("Error : {} {} ".format(schemaname,tablename))
      print("select distinct loaddate from {0}.{1} where loaddate <={2} order by loaddate limit 2".format(schemaname,tablename,loaddate))
      print(str(e))
      return {"sql error" :"sql error","loaddate":"error"}
    SecondMax_loaddate = loaddate

  if column_names :
      # building the select string 
      sql_string = "select "
      for idx,col_names in enumerate(column_names.split(',')) :
        sql_string +=" max("+col_names+") as "+ col_names + ( "," if idx != len(column_names.split(','))-1 else "")
      sql_string +=" from {0}.{1} where loaddate ='{2}'".format(schemaname,tablename,SecondMax_loaddate)
      # print(sql_string)
      try:
        hive_curs.execute(sql_string)        
      except Exception as e:
        print("Error : {0}.{1}".format(schemaname,tablename))
        print(sql_string)
        print(str(e))
        return {"sql error ": "sql error","loaddate":loaddate_rows[0][0]}

      return_dict = getRowsAsDict(hive_curs)[0]
      return_dict['loaddate'] = loaddate_rows[0][0]
      return return_dict

  else :
    # partition -1 max date should be updated -need to check
    hive_curs.execute("desc {}.{}".format(schemaname,tablename))    
    resultset = getRowsAsDict(hive_curs)
    #removing the partition section from the describe command

    try:
        target_index = resultset.index({u'col_name': u'', u'comment': None, u'data_type': None})
    except ValueError, e:
        target_index = None
    # {u'col_name': u'dw_batch_id', u'comment': u'', u'data_type': u'int'},   

    resultset = resultset[:target_index]
    cols_available  = [dicts['col_name'] for dicts in resultset]
    cols_available = map(lambda x: x.lower().strip(),cols_available)
    time_cols_avl = set(cols_available).intersection(set(['activationdatekey','dateadded','datecreated','datemodified','deactivationdatekey','dw_create_dt','dw_insert_dt','dw_last_update_dt','remove_event_dt','remove_event_tm']))


    if time_cols_avl :
      # building the select string 
      sql_string = "select "
      for idx,col_names in enumerate(time_cols_avl) :
        sql_string +=" max("+col_names+") as "+ col_names + ( "," if idx != len(time_cols_avl)-1 else "")

      sql_string+=" from {0}.{1} where loaddate ='{2}'".format(schemaname,tablename,SecondMax_loaddate)

      # print(sql_string)
      try:
        hive_curs.execute(sql_string)
        return_val = getRowsAsDict(hive_curs)
      except Exception as e:
        print("Error : {0}.{1}".format(schemaname,tablename))
        print(sql_string)
        print(str(e))
        return {"sql error ": "sql error","loaddate":loaddate_rows[0][0]}     
      
      # just return the first row
      # print(return_val[0])
      return_val[0]['loaddate'] = loaddate_rows[0][0]
      return return_val[0]
    else :
      #there no timestamp columns in this table to do CDC so a fullload table.
      return {"No CDC columns found":"No CDC columns found","loaddate":loaddate_rows[0][0]}

def do_SZcountvalidation() :
  print "\n=========================== SZ Count Validation =====================================\n";
  while True:
    InputFileName = raw_input(" Enter the input parameter file path with filename (.csv): ")
    ParasFilePath = InputFileName
    if ParasFilePath == "":
      print " Please provide the Input parameter file !!!"
      continue
    else:
      break

  try:
      f_lock = FileLock(ParasFilePath)
      f = open(ParasFilePath, 'r')
  except IOError:
      print "Could not read file:", ParasFilePath
      sys.exit()

  try:
      OutputFilename = OutputFilePath+"SZ_count_validation"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
      outputfile = open(OutputFilename, 'w')
  except IOError:
      print "Could not read file:", OutputFilename
      sys.exit()

  print "\n============================== Execution In-progress ============================================\n";
  print ' Record Count Validation started at ' + str(datetime.datetime.now())
  ExecuStartTime = str(datetime.datetime.now()) 

  initialize_hive()    
  with f_lock :
    with f:
      outputfile.write(
          "Id,\
          Src-Schema,\
          Src_table,\
          SZ-Schema,\
          SZ-table,\
          Src-RecordCount,\
          SZ-RecordCount,\
          status\n")
      print(
          "Id,\
          Src-Schema,\
          Src_table,\
          SZ-Schema,\
          SZ-table,\
          Src-RecordCount,\
          SZ-RecordCount,\
          status")

      reader = csv.DictReader(f)
      count = 1
      passed,failed = 0,0
      for row in reader : 
          src_count = gethiveszsqlcount(row['Src_Sql'])
        # if row['Sz_Sql'] == '' or len(row['Sz_Sql']) == 0 or row['Sz_Sql'].strip() == '':
        #   sz_count = gethiveszcount(row['Sz_SchemaName'],row['Sz_HiveTableName'])
        # else :
          sz_count = gethiveszsqlcount(row['Sz_Sql'])

      if src_count == sz_count :
        match_mismatch = "Match"
      else :
        match_mismatch = "MisMatch"


      outputfile.write(
      str(count)
      +","+row['Src_SchemaName'] 
      +","+row['Src_table']
      +","+row['Sz_SchemaName']
      +","+row['Sz_HiveTableName']
      # +","+row['Sz_Sql']
      +","+str(src_count)
      +","+str(sz_count)
      +","+match_mismatch
      +"\n" )

      print(str(count)
      +","+row['Src_SchemaName'] 
      +","+row['Src_table']
      +","+row['Sz_SchemaName']
      +","+row['Sz_HiveTableName']
      # +","+row['Sz_Sql']
      +","+str(src_count)
      +","+str(sz_count)
      +","+match_mismatch
      )

      count +=1
      if match_mismatch == "Match" :
        passed += 1;
      else :
        failed += 1;
          
    
    validation = "SZ Count"
    ostring = printsummary(validation,count-1,passed,failed,ParasFilePath,OutputFilename,ExecuStartTime)
    outputfile.write(ostring)
    print(ostring)
    

    outputfile.close()
    close_hive()


def do_CDCUpdate() :
  print "\n=========================== CDC Update =====================================\n";

  # Environment = 0
  # while True:
  #   Environment = raw_input(" Interim Dev (1)\n Strategic Dev (2) \n Strategic Test (3) \n Pre-Prod (4) \n Prod (5) \n Enter the Hive environment Number: ")
  #   if Environment == "" or len(Environment) != 1 or Environment.isalpha():
  #     print " Please provide the valid Environment Number !!!"
  #     continue
  #   else:
  #     break


  while True:
    InputFileName = raw_input(" Enter the input parameter file path with filename (.xlsx): ")
    ParasFilePath = InputFileName
    if ParasFilePath == "":
      print " Please provide the Input parameter file !!!"
      continue
    else:
      break

  try:
      f_lock = FileLock(ParasFilePath)
      f = open(ParasFilePath, 'r')
  except IOError:
      print "Could not read file:", ParasFilePath
      sys.exit()

  try:
      OutputFilename = OutputFilePath+"CDC_InputFile_Update"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
      outputfile = open(OutputFilename, 'w')
  except IOError:
      print "Could not read file:", OutputFilename
      sys.exit()
  print "\n============================== CDC Update In-progress ============================================\n";
  print ' CDC Update started at ' + str(datetime.datetime.now())
  ExecuStartTime = str(datetime.datetime.now()) 

  initialize_hive()    
  with f_lock :
    with f:
        from tempfile import NamedTemporaryFile
        import shutil

        tempfile = NamedTemporaryFile(mode='w',delete=False)
        reader = XLSXDictReader1(f)
        from openpyxl import Workbook,load_workbook
        # temp_wb = Workbook()
        # temp_sheet = temp_wb.create_sheet("input",0)

        wb = load_workbook(f.name) 

        # removing the exisintg input sheet
        # input_sheet=wb.get_sheet_by_name('input')
        wb.remove(wb['Input'])

        # creating a new input sheet
        temp_sheet = wb.create_sheet("Input",0)        
        temp_sheet.append(reader[0])

        # writer = csv.DictWriter(tempfile, fieldnames=reader[0])
        # writer.writeheader()
        count = 1
        for row in reader[1] :
          # print(row)
          # if cdc dates are missing then do the update for that row
          if not row['HiveTableName'] is None :
            count += 1
            if row['Loadtype'].lower().strip() != 'historical'  or row['Loadtype'].lower().strip()  != 'incremental-full' :
              if row['CDC-Dates'] is None or row['CDC-Dates'].strip() == '' or len(row['CDC-Dates']) == 0 :
                if row['CDC-Columns'] == '' or len(str(row['CDC-Columns'])) == 0 or str(row['CDC-Columns']).strip() == '':
                  if row['LoadDate'] == '' or len(str(row['LoadDate'])) == 0 or str(row['LoadDate']).strip() == '':
                    columns_timestamps = getcdchive(row['LZ_SchemaName'],row['HiveTableName'])
                  else :
                    columns_timestamps = getcdchive(row['LZ_SchemaName'],row['HiveTableName'],loaddate=row['LoadDate'])

                  row['LoadDate'] = columns_timestamps.pop('loaddate')
                  row['CDC-Columns'] = ','.join(columns_timestamps.keys())
                  row['CDC-Dates'] = ','.join(map(str,columns_timestamps.values()))
                  # row['CDC-Dates'], row['CDC-Columns']= timestamps.get('dw_create_dt',timestamps.get('datecreated','')), timestamps.get('dw_last_update_dt',timestamps.get('datemodified',''))

                else : 
                  if row['LoadDate'] == '' or len(str(row['LoadDate'])) == 0 or str(row['LoadDate']).strip() == '':
                    columns_timestamps = getcdchive(row['LZ_SchemaName'],row['HiveTableName'],column_names=row['CDC-Columns'])
                  else :
                    columns_timestamps = getcdchive(row['LZ_SchemaName'],row['HiveTableName'],loaddate=row['LoadDate'],column_names=row['CDC-Columns'])             
                  # timestamps = getcdchive(row['LZ_SchemaName'],row['HiveTableName'],row['CDC-Columns']) # these come preformatted ie comma seperated.
                  row['LoadDate'] = columns_timestamps.pop('loaddate')
                  row['CDC-Dates'] = ','.join(columns_timestamps.values())
              # CDC-Dates CDC-Columns
              outputfile.write('updating row : '+str(row['ID']) +" schemaname : "+row['LZ_SchemaName'] +" tablename : "+row['HiveTableName'])
              print('updating row : '+str(row['ID']) +" schemaname : "+row['LZ_SchemaName'] +" tablename : "+row['HiveTableName'])
          temp_sheet.append(row.values())
          # writer.writerow(row)
  print("\n================================ CDC Update Summary===============================================\n")
  print(" Total number of Tables Updated                      : " + str(count))
  print("\n")
  print(" Input parameter file        : " + ParasFilePath)
  # print(" Raw file contains header?   : " + Header)
  print(" Output File has been stored : " + OutputFilename)
  print(" Environment selected        : " + os.environ['env'] if os.environ['env'] else os.environ['ENV'])
  print(" Execution started Time      : " + ExecuStartTime)
  print(" Execution End Time          : " + str(datetime.datetime.now()))
  print("\n================================================================================================\n")


  outputfile.close()
  close_hive()
  # print(tempfile.name)
  wb.save(f.name)
  # shutil.move(tempfile.name, ParasFilePath)

def do_CountValidation(loadtype) :
  print "\n=========================== Record Count Validation Inputs =====================================\n";
  while True:
    InputFileName = raw_input(" Enter the input parameter file path with filename (.xlsx): ")
    #InputFileName = "/home/xwislg6/Automation/input.csv"
    ParasFilePath = InputFileName
    if ParasFilePath == "":
      print " Please provide the Input parameter file !!!"
      continue
    else:
      break

  #Header = raw_input(" Does the raw file contains header record? (Yes/No): ")
  Header = "No"
  if Header == "":
    Header = "No"
    print " Default Value Assigned: No Header record in the file"
  #OutputFilePath = raw_input(" Enter the output file path only: ")

  if OutputFilePath == "":
    print " By default Output file will be stored in your current folder !!!"

  # Environment = 0
  # while True:
  #   Environment = raw_input(" Interim Dev (1)\n Strategic Dev (2) \n Strategic Test (3) \n Pre-Prod (4) \n Prod (5) \n Enter the Hive environment Number: ")
  #   if Environment == "" or len(Environment) != 1 or Environment.isalpha():
  #     print " Please provide the valid Environment Number !!!"
  #     continue
  #   else:
  #     break
  #Environment = "1"

  print "\n============================== Execution In-progress ============================================\n";
  print ' Record Count Validation started at ' + str(datetime.datetime.now())
  ExecuStartTime = str(datetime.datetime.now()) 

  try:
      f_lock = FileLock(ParasFilePath)
      f = open(ParasFilePath, 'rb')
  except IOError:
      print "Could not read file:", ParasFilePath
      sys.exit()

  try:
      OutputFilename = OutputFilePath+"LZ_OZ_"+loadtype+"_Count_ABC_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"  
      outputfile = open(OutputFilename, 'w')
  except IOError:
      print "Could not read file:", OutputFilename
      sys.exit()

  initialize_hive()
  initialize_abcmssql()
  with f_lock :
    with f:
        reader = XLSXDictReader(f)
        templist = list(reader[1])

        #checkFileValid(templist)

        # sortedreader = sorted(tempreader, key=lambda row:(row['S_Server'],row['S_Port'])),row['S_Database']
        from itertools import groupby
        srcservergrouped = groupby(templist,key=lambda row:(row['DatabaseType'],row['S_Server']))
        with outputfile :
          outputfile.write(
          "Id, \
          Source DB, \
          SourceSchema, \
          Sourcetable, \
          LoadType, \
          LoadDate, \
          CDC-Dates, \
          Source-RecordCount, \
          Source-FullRecordCount, \
          LZ-RecordCount, \
          OZ-RecordCount, \
          status, \
          LZ_batch_id,\
          LZ_ABC_src_count, \
          LZ_ABC_tgt_count, \
          LZ_ABC_Status,\
          LZ_balanced_ind,\
          OZ_batch_id,\
          OZ_ABC_src_count, \
          OZ_ABC_tgt_count, \
          OZ_ABC_Status,\
          OZ_balanced_ind\
          \n")

          print(
          "Id, \
          Source DB, \
          SourceSchema, \
          Sourcetable, \
          LoadType, \
          LoadDate, \
          CDC-Dates, \
          Source-RecordCount, \
          Source-FullRecordCount, \
          LZ-RecordCount, \
          OZ-RecordCount, \
          status, \
          LZ_batch_id,\
          LZ_ABC_src_count, \
          LZ_ABC_tgt_count, \
          LZ_ABC_Status,\
          LZ_balanced_ind,\
          OZ_batch_id,\
          OZ_ABC_src_count, \
          OZ_ABC_tgt_count, \
          OZ_ABC_Status,\
          OZ_balanced_ind\
          ")
          count = 1
          passed = 0;
          failed = 0;  
          abc_status =''
          
          for key,rows in srcservergrouped:
              if key[0] == 'MSSQL' :
                #connect to sqlserver for each source by passing the server name found in windows
                initialize_mssql(key[1])
              elif key[0] == 'Netezza' :
                #connect to neteeza for each source by passing the server name found in windows
                initialize_neteeza(key[1])

              for row in rows :
                if  row['ACTIVE'] != 'Y' :
                  continue

                full_record_src_count = 0

                if key[0] == 'Netezza' :
                  # if the loadtype is Historical or Incremental-full then we dont get the count based on create and update date

                  if loadtype == "hist" :
                    # and ( row['Loadtype'].lower().strip() == 'historical'  or row['Loadtype'].lower().strip()  == 'incremental-full') : if it is historical or firt load then get all the counts as full
                    src_count = getneteezarowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],checkCDC=False) 
                  # else we do the the Count based on create and update date
                  elif loadtype == "incremental" and ( row['Loadtype'].lower().strip() == 'historical'  or row['Loadtype'].lower().strip()  == 'incremental-full') :
                    src_count = getneteezarowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],checkCDC=False) 
                  elif loadtype == "incremental" and row['Loadtype'].lower().strip()  == 'incremental-cdc' :
                    full_record_src_count = getneteezarowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],checkCDC=False)
                    src_count = getneteezarowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],CDC_Columns=row['CDC-Columns'],CDC_Dates=row['CDC-Dates'],checkCDC=True)

                elif key[0] == 'MSSQL' :    
                  if loadtype == "hist" :
                    src_count = getsqlrowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],checkCDC=False) 
                  # if the loadtype is Historical or Incremental-full then we dont get the count based on create and update date
                  elif loadtype == "incremental" and ( row['Loadtype'].lower().strip() == 'historical'  or row['Loadtype'].lower().strip()  == 'incremental-full') :
                    src_count = getsqlrowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],checkCDC=False) 
                  # else we do the the Count based on create and update date
                  elif loadtype == "incremental" :
                    full_record_src_count = getsqlrowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],checkCDC=False)
                    src_count = getsqlrowcount(row['S_Database'],row['S_Schema'],row['S_TableName'],CDC_Columns=row['CDC-Columns'],CDC_Dates=row['CDC-Dates'],checkCDC=True)                  

                skipOZ = False
                skipSZ = False

                if row['Oz_SchemaName'] == '' or len(row['Oz_SchemaName']) == 0 or row['Oz_SchemaName'].strip() == '':
                    skipOZ = True

                # if row['SZ_HiveTableName'] == '' or len(row['SZ_HiveTableName']) == 0 or row['SZ_HiveTableName'].strip() == '':
                #     skipSZ = True
                
                if row['LZ_BatchId'] :
                  abc_status = getABC_Status(row['LZ_BatchId'])
                else :
                  abc_status = getABC_StatusTableName(row['S_TableName'])


                if row['OZ_BatchId'] :
                  abc_status = abc_status + getABC_Status(row['OZ_BatchId'])
                else :
                  abc_status = getABC_StatusTableName(row['S_TableName'])
                  



                LZ_hive_count = ""
                OZ_hive_count = ""
                # SZ_hive_count = ""

                # check partition if it is null string or empty then return Null
                if row['Loadtype'].lower().strip() == 'historical' :
                  hive_lz_result = gethivelzcount(row['LZ_SchemaName'],row['HiveTableName'],( None if str(row['LoadDate']).strip() =='' or str(row['LoadDate']) is None else str(row['LoadDate']).strip() ) )
                else :
                  hive_lz_result = gethiveozcount(row['LZ_SchemaName'],row['HiveTableName'])  #using oz count so that we only run count(*)
                  hive_lz_result = [ '' if str(row['LoadDate']).strip() =='' or str(row['LoadDate']) is None else str(row['LoadDate']).strip(),hive_lz_result]

                if isinstance(hive_lz_result,int) or isinstance(hive_lz_result,jpype.java.lang.Long):
                  LZ_hive_count = hive_lz_result                  
                else :
                  print(hive_lz_result)
                  LZ_hive_count = hive_lz_result[1]
                  row['LoadDate'] = hive_lz_result[0]

                if not skipOZ :
                  OZ_hive_count = gethiveozcount(row['Oz_SchemaName'],row['HiveTableName'])

                # if not skipSZ :
                #   sql = row['SZ_SQL'] if row['SZ_SQL'] else "select count(*) as count from {}.{}".format(row['Sz_SchemaName'],row['SZ_HiveTableName'])
                #   SZ_hive_count = gethiveszsqlcount(sql)
                if loadtype != "hist" :
                # if row['Loadtype'].lower().strip() != 'historical' and row['Loadtype'].lower().strip()  != 'incremental-full' :
                  if not skipOZ :
                    # if not skipSZ :
                    #   if src_count == LZ_hive_count and LZ_hive_count == OZ_hive_count  and OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "Match"
                    #   elif src_count == LZ_hive_count and LZ_hive_count == OZ_hive_count :
                    #     match_mismatch = "OZ - SZ Match"
                    #   elif src_count == LZ_hive_count and OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "lZ - OZ Match"
                    #   elif LZ_hive_count == OZ_hive_count and OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "Source - LZ MisMatch"
                    #   elif src_count == LZ_hive_count :
                    #     match_mismatch = "LZ - OZ - SZ MisMatch"
                    #   elif LZ_hive_count == OZ_hive_count :
                    #     match_mismatch = "Source - LZ - SZ MisMatch"
                    #   elif OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "Source - LZ - OZ MisMatch"
                    #   else :
                    #     match_mismatch = "All Counts MisMatch"

                    # else :
                      if src_count == LZ_hive_count and full_record_src_count == OZ_hive_count :
                        match_mismatch = "Match"
                      elif src_count == LZ_hive_count :
                        match_mismatch = "Source - OZ MisMatch"
                      elif full_record_src_count == OZ_hive_count :
                        match_mismatch = "Source - LZ MisMatch"
                      else :
                        match_mismatch = "All Counts MisMatch"
                  else :
                    if src_count == LZ_hive_count :
                      match_mismatch = "Match"
                    else :
                      match_mismatch = "MisMatch"
                else :
                  if not skipOZ :
                    # if not skipSZ :
                    #   if src_count == LZ_hive_count and LZ_hive_count == OZ_hive_count  and OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "Match"
                    #   elif src_count == LZ_hive_count and LZ_hive_count == OZ_hive_count :
                    #     match_mismatch = "OZ - SZ Match"
                    #   elif src_count == LZ_hive_count and OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "lZ - OZ Match"
                    #   elif LZ_hive_count == OZ_hive_count and OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "Source - LZ MisMatch"
                    #   elif src_count == LZ_hive_count :
                    #     match_mismatch = "LZ - OZ - SZ MisMatch"
                    #   elif LZ_hive_count == OZ_hive_count :
                    #     match_mismatch = "Source - LZ - SZ MisMatch"
                    #   elif OZ_hive_count == SZ_hive_count :
                    #     match_mismatch = "Source - LZ - OZ MisMatch"
                    #   else :
                    #     match_mismatch = "All Counts MisMatch"

                    # else :
                      print("src_count >{}< LZ_hive_count >{}< OZ_hive_count >{}<".format(src_count,LZ_hive_count,OZ_hive_count))
                      if src_count == LZ_hive_count and LZ_hive_count == OZ_hive_count :
                        match_mismatch = "Match"
                      elif src_count == LZ_hive_count :
                        match_mismatch = "LZ - OZ MisMatch"
                      elif LZ_hive_count == OZ_hive_count :
                        match_mismatch = "Source - LZ MisMatch"
                      else :
                        match_mismatch = "All Counts MisMatch"
                  else :
                    if src_count == LZ_hive_count :
                      match_mismatch = "Match"
                    else :
                      match_mismatch = "MisMatch"




                outputfile.write( \
                "\""+str(count) +"\""\
                +",\""+row['S_Database']  +"\""\
                +",\""+row['S_Schema'] +"\""\
                +",\""+row['S_TableName'] +"\""\
                +",\""+("" if row['LoadDate'] is None or row['LoadDate'] == "None" else str(row['Loadtype'])) +"\""\
                #+",\""+str(row['LoadDate'])  +"\""\
                # +",\""+row['CDC-Dates'] +"\""\
                +",\""+("" if row['CDC-Dates'] is None or row['CDC-Dates'] == "None" else str(row['CDC-Dates']))+"\""\
                +",\""+str(src_count) +"\""\
                +",\""+str(full_record_src_count) +"\""\
                +",\""+str(LZ_hive_count) +"\""\
                +",\""+str(OZ_hive_count) +"\""\
                # +",\""+str(SZ_hive_count) +"\""\
                +",\""+match_mismatch +"\""\
                +","+abc_status
                # +","+str(abc_status['batch_set_key'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_key'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_set_src_sys_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_set_category'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_set_sts_descr'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_src_db_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_trgt_db_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_sts_descr'] if not abc_status is None else "") \
                # +","+str(abc_status['measure_type'] if not abc_status is None else "") \
                # +","+str(abc_status['measure_src_value'] if not abc_status is None else "") \
                # +","+str(abc_status['measure_trgt_value'] if not abc_status is None else "") \
                # +","+str(abc_status['balanced_ind'] if not abc_status is None else "") \
                +"\n" )

                print( \
                "\""+str(count) +"\""\
                +",\""+row['S_Database'] +"\""  \
                +",\""+row['S_Schema'] +"\"" \
                +",\""+row['S_TableName'] +"\"" \
                +",\""+("" if row['Loadtype'] is None or row['Loadtype'] == "None" else str(row['Loadtype']))+"\""\
                #+",\""+row['Loadtype'] +"\"" \
                +",\""+str(row['LoadDate'])  +"\"" \
                +",\""+("" if row['CDC-Dates'] is None or row['CDC-Dates'] == "None" else str(row['CDC-Dates']))+"\""\
                +",\""+str(src_count) +"\"" \
                +",\""+str(full_record_src_count) +"\"" \
                +",\""+str(LZ_hive_count) +"\"" \
                +",\""+str(OZ_hive_count) +"\"" \
                # +",\""+str(SZ_hive_count) +"\"" \
                +",\""+match_mismatch +"\"" \
                +","+abc_status \
                # +","+str(abc_status['batch_key'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_set_src_sys_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_set_category'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_set_sts_descr'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_src_db_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_trgt_db_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_nme'] if not abc_status is None else "") \
                # +","+str(abc_status['batch_sts_descr'] if not abc_status is None else "") \
                # +","+str(abc_status['measure_type'] if not abc_status is None else "") \
                # +","+str(abc_status['measure_src_value'] if not abc_status is None else "") \
                # +","+str(abc_status['measure_trgt_value'] if not abc_status is None else "") \
                # +","+str(abc_status['balanced_ind'] if not abc_status is None else "") 
                )
                # print(str(count)+" SourceSchema:{} Sourcetable:{} TargetSchema:{} Targettable:{} sourcecount:{} targetcount:{} status:{}".format(row['S_Schema'],row['S_TableName'],row['LZ_SchemaName'],row['HiveTableName'],str(sql_count),str(hive_count),match_mismatch))
                count +=1
                if match_mismatch == "Match" :
                  passed += 1;
                else :
                  failed += 1;

              close_mssql()

          
          if loadtype == "hist":

            validation = "LZ and OZ Hist Count"
          else :
            validation = "LZ and OZ Incremental Count"
          ostring = printsummary(validation,count-1,passed,failed,ParasFilePath,OutputFilename,ExecuStartTime)
          outputfile.write(ostring)
          print(ostring)


    close_abcmssql()
    close_hive()
    return

def do_SZMetadataValidation() :
  print "\n=========================== SZ Metadata Validation =====================================\n";
  while True:
    InputFileName = raw_input(" Enter the input parameter file path with filename (.xlsx): ")
    ParasFilePath = InputFileName
    if ParasFilePath == "":
      print " Please provide the Input parameter file !!!"
      continue
    else:
      break

  try:
      f_lock = FileLock(ParasFilePath)
      f = open(ParasFilePath, 'r')
  except IOError:
      print "Could not read file:", ParasFilePath
      sys.exit()

  # Environment = 0
  # while True:
  #   Environment = raw_input(" Interim Dev (1)\n Strategic Dev (2) \n Strategic Test (3) \n Pre-Prod (4) \n Prod (5) \n Enter the Hive environment Number: ")
  #   if Environment == "" or len(Environment) != 1 or Environment.isalpha():
  #     print " Please provide the valid Environment Number !!!"
  #     continue
  #   else:
  #     break  

  print "\n============================== SZ Metadata Validation Execution In-progress ============================================\n";
  print ' SZ Metadata Validation started at ' + str(datetime.datetime.now())
  ExecuStartTime = str(datetime.datetime.now()) 

  # initialize_hivemeta('acenmetasqlserver01d')
  initialize_hive()
  initialize_datamodelconversion()

  try:
    src_sz_MetaOutputFilename = "../output/"+"SRC_SZ_MetadataValidation_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
    sz_datamodel_MetaOutputFilename = "../output/"+"SZ_DATAMODEL_MetadataValidation_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
    # MetaOutputFilePath+row['S_TableName']+
    src_sz_outputfile = open(src_sz_MetaOutputFilename, 'w')
    datamodel_sz_outputfile = open(sz_datamodel_MetaOutputFilename, 'w')
  except IOError:
    print "Could not read file:", sz_datamodel_MetaOutputFilename
    sys.exit()

  src_sz_outputfile.write("id,src,SchenamName,TableName,ColumnName,OriginalDatatype,Datatype,Column-index"
  # +" ,"+",lz SchenamName,lz TableName,lz ColumnName,lz OriginalDatatype,lz Datatype,lz Column-index,src-lz Datatype-Status,src-lz Column-Status"
  # +" "+",oz SchenamName,oz TableName,oz ColumnName,oz OriginalDatatype,oz Datatype,oz Column-index,lz-oz Datatype-Status,lz-oz Column-Status"
  +" ,"+",sz SchenamName,sz TableName,sz ColumnName,sz OriginalDatatype,sz Datatype,sz Column-index,src-sz Datatype-Status,src-sz Column-Status"
  +"\n")

  datamodel_sz_outputfile.write("id,src,SchenamName,Datamodel TableName,Datamodel ColumnName,Datamodel OriginalDatatype,Datamodel Datatype,Datamodel Column-index"
  # +" ,"+",oz SchenamName,oz TableName,oz ColumnName,oz OriginalDatatype,oz Datatype,oz Column-index,lz-oz Datatype-Status,lz-oz Column-Status"
  +" ,"+",sz SchenamName,sz TableName,sz ColumnName,sz OriginalDatatype,sz Datatype,sz Column-index,Datamodel-sz Datatype-Status,Datamodel-sz Column-Status"
  +"\n")

  with f_lock :
    with f:
      reader = XLSXDictReader(f)
      templist = list(reader[1])

      from itertools import groupby
      srcservergrouped = groupby(templist,key=lambda row:(row['DatabaseType'],row['S_Server']))

      for key,rows in srcservergrouped :
        if key[0] =='MSSQL' :
          initialize_mssql(key[1]) 
        
        srcsz_failed,srcsz_passed = 0,0
        datamdl_failed,datamdl_passed = 0,0
        count,passed,failed = 0,0,0
        for row in rows :

          if key[0] =='MSSQL' :
            # print(row['S_TableName'])
            if not ( row['S_TableName'] == '' or len(row['S_TableName']) == 0 or row['S_TableName'].strip() == ''):
              mssql_columns_records = get_mssql_structure(row['S_Database'],row['S_Schema'],row['S_TableName'])
              src_lower_cols = transform_sqldatatype(lowercase_columns(mssql_columns_records))

          elif key[0] == 'Netezza':
            # print(row['S_TableName'])
            neteeza_columns_records = get_neteeza_structure(row['S_TableName'],neteeza_curs)
            src_lower_cols = lowercase_columns(neteeza_columns_records)
          # print(row['SZ_SchemaName'])
          if (row['SZ_SchemaName'] is None or row['SZ_SchemaName'] == 'None'):
            continue;
          else :
            # print(row['SZ_SchemaName'] + 'check 2')
            sz_hive_columns_records = get_hive_structure_v2(row['SZ_SchemaName'],row['HiveTableName'],hive_curs) 
            sz_hive_lower_cols = lowercase_columns(sz_hive_columns_records)

            # if table missing just add table is missing          
            datamodel_hive_columns_records = get_columnsFromDatamodel(row['HiveTableName'])
          if datamodel_hive_columns_records :            
            datamodel_hive_lower_cols = lowercase_columns(datamodel_hive_columns_records)
          else :
            print("Table not present in Datamodel : "+row['HiveTableName'])


          if not (row['S_TableName'] == '' or len(row['S_TableName']) == 0 or row['S_TableName'].strip() == '' ):
            src_sz_compare = do_comparison(sz_hive_lower_cols,src_lower_cols,row['Source'])
            src_sz_outputfile.write(src_sz_compare)

            if datamodel_hive_columns_records :
              datamodel_sz_compare = do_comparison(datamodel_hive_lower_cols,sz_hive_lower_cols,row['Source'])
              datamodel_sz_outputfile.write(datamodel_sz_compare)
            # oz_sz_compare = do_comparison(sql_lower_cols,sz_hive_lower_cols)
            # src_sz_outputfile.write(re.sub(r'^',row['Source'],src_lz_compare))
            # datamodel_sz_outputfile.write(re.sub(r'^',row['Source'],lz_oz_compare))
            count += 1
            if "MisMatch"  in src_sz_compare or "missing"  in src_sz_compare :
              srcsz_failed += 1
            else :
              srcsz_passed += 1

            if "MisMatch"  in src_sz_compare or "missing"  in datamodel_sz_compare :
              datamdl_failed += 1
            else :
              datamdl_passed += 1


            
            

            # row['SZ_HiveTableName'] == '' or len(row['SZ_HiveTableName']) == 0 or row['SZ_HiveTableName']
            # row['Sz_SchemaName'],row['SZ_HiveTableName']
            
            print(row['S_TableName']+": Completed")

      src_sz_outputfile.write("\n================================Execution Summary===============================================\n")
      src_sz_outputfile.write(" Total number of Tables Validated                    : " + str(count)+"\n")
      src_sz_outputfile.write(" Number of Tables passed Metadata Validation : " + str(srcsz_passed)+"\n")
      src_sz_outputfile.write(" Number of Tables Failed Metadata Validation : " + str(srcsz_failed)+"\n")
      src_sz_outputfile.write("\n")
      src_sz_outputfile.write(" Input parameter file        : " + ParasFilePath+"\n")
      # src_sz_outputfile.write(" Raw file contains header?   : " + Header)
      src_sz_outputfile.write(" Output File has been stored : " + src_sz_MetaOutputFilename +"\n")
      src_sz_outputfile.write(" Environment selected        : " + Environment+"\n") #change to environment name
      src_sz_outputfile.write(" Execution started Time      : " + ExecuStartTime+"\n")
      src_sz_outputfile.write(" Execution End Time          : " + str(datetime.datetime.now())+"\n")
      src_sz_outputfile.write("\n================================================================================================\n")

      datamodel_sz_outputfile.write("\n================================Execution Summary===============================================\n")
      datamodel_sz_outputfile.write(" Total number of Tables Validated                    : " + str(count)+"\n")
      datamodel_sz_outputfile.write(" Number of Tables passed Metadata Validation : " + str(datamdl_passed)+"\n")
      datamodel_sz_outputfile.write(" Number of Tables Failed Metadata Validation : " + str(datamdl_failed)+"\n")
      datamodel_sz_outputfile.write("\n")
      datamodel_sz_outputfile.write(" Input parameter file        : " + ParasFilePath+"\n")
      # datamodel_sz_outputfile.write(" Raw file contains header?   : " + Header)
      datamodel_sz_outputfile.write(" Output File has been stored : " + sz_datamodel_MetaOutputFilename +"\n")
      datamodel_sz_outputfile.write(" Environment selected        : " + Environment+"\n") #change to environment name
      datamodel_sz_outputfile.write(" Execution started Time      : " + ExecuStartTime+"\n")
      datamodel_sz_outputfile.write(" Execution End Time          : " + str(datetime.datetime.now())+"\n")
      datamodel_sz_outputfile.write("\n================================================================================================\n")
      #validation = "SZ Metadata"
      #printsummary(validation,count-1,passed,failed,ParasFilePath,src_sz_MetaOutputFilename+" "+sz_datamodel_MetaOutputFilename,ExecuStartTime)
     
      print("\n================================ Execution Summary ===============================================\n")
      print(" Total number of Tables Validated                    : " + str(count))
      print(" Number of Tables passed SRC SZ Metadata Validation : " + str(srcsz_passed))
      print(" Number of Tables Failed SRC SZ Metadata Validation : " + str(srcsz_failed))
      print(" Number of Tables passed Datamodel SZ Metadata Validation : " + str(datamdl_passed))
      print(" Number of Tables Failed Datamodel SZ Metadata Validation : " + str(datamdl_failed))
      print("\n")
      print(" Input parameter file        : " + ParasFilePath)
      # print(" Raw file contains header?   : " + Header)
      print(" Output File has been stored : " + src_sz_MetaOutputFilename+" "+sz_datamodel_MetaOutputFilename )
      print(" Environment selected        : " + Environment)
      print(" Execution started Time      : " + ExecuStartTime)
      print(" Execution End Time          : " + str(datetime.datetime.now()))
      print("\n================================================================================================\n")

      # close_hivemeta()
      close_mssql()
      src_sz_outputfile.close()
      datamodel_sz_outputfile.close()

def do_AllValidationHist() :


  STATMain(1) 
  do_CountValidation("hist")
  do_MetadataValidation()
  STATMain(2)
  log_validation("Landing")
  sz_count_validation()
  do_SZMetadataValidation()
  sz_azure_count_validation()
  sz_azure_metadata_validation()
  STATMain(3)
  do_denodocount()

def do_AllValidationIncr() :
  # do_CDCUpdate()
  #do_CountValidation("hist")
  STATMain(1)
  do_CountValidation("incremental")
  do_MetadataValidation()
  STATMain(2)
  sz_count_validation()
  do_SZMetadataValidation()
  sz_azure_count_validation()
  sz_azure_metadata_validation()
  STATMain(3)
  do_denodocount()
  #do_DenodoCountValidation()
    
def do_MetadataValidation() :
  print "\n=========================== Metadata Validation =====================================\n";
  while True:
    InputFileName = raw_input(" Enter the input parameter file path with filename (.xlsx): ")
    ParasFilePath = InputFileName
    if ParasFilePath == "":
      print " Please provide the Input parameter file !!!"
      continue
    else:
      break

  try:
      f_lock = FileLock(ParasFilePath)
      f = open(ParasFilePath, 'r')
  except IOError:
      print "Could not read file:", ParasFilePath
      sys.exit()

  # Environment = 0
  # while True:
  #   Environment = raw_input(" Interim Dev (1)\n Strategic Dev (2) \n Strategic Test (3) \n Pre-Prod (4) \n Prod (5) \n Enter the Hive environment Number: ")
  #   if Environment == "" or len(Environment) != 1 or Environment.isalpha():
  #     print " Please provide the valid Environment Number !!!"
  #     continue
  #   else:
  #     break
  
  # initialize_hivemeta('acenmetasqlserver01d')
  initialize_hive()
  initialize_neteeza()
  try:
    src_lz_MetaOutputFilename = "../output/"+"SRC_LZ_MetadataValidation_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
    lz_oz_MetaOutputFilename = "../output/"+"LZ_OZ_MetadataValidation_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
    # MetaOutputFilePath+row['S_TableName']+
    src_lz_outputfile = open(src_lz_MetaOutputFilename, 'w')
    lz_oz_outputfile = open(lz_oz_MetaOutputFilename, 'w')
  except IOError:
    print "Could not read file:", MetaOutputFilename
    sys.exit()
  print "\n============================== LZ OZ Metadata Validation In-progress ============================================\n";
  print ' LZ OZ Metadata Validation started at ' + str(datetime.datetime.now())
  ExecuStartTime = str(datetime.datetime.now()) 
  src_lz_outputfile.write("id,src,SchenamName,TableName,ColumnName,OriginalDatatype,Datatype,Column-index"
  +" ,"+",lz SchenamName,lz TableName,lz ColumnName,lz OriginalDatatype,lz Datatype,lz Column-index,src-lz Datatype-Status,src-lz Column-Status"
  # +" "+",oz SchenamName,oz TableName,oz ColumnName,oz OriginalDatatype,oz Datatype,oz Column-index,lz-oz Datatype-Status,lz-oz Column-Status"
  # +" "+",sz SchenamName,sz TableName,sz ColumnName,sz OriginalDatatype,sz Datatype,sz Column-index,src-sz Datatype-Status,src-sz Column-Status"
  +"\n")

  lz_oz_outputfile.write("id,src,lz SchenamName,lz TableName,lz ColumnName,lz OriginalDatatype,lz Datatype,lz Column-index"
  +" ,"+",oz SchenamName,oz TableName,oz ColumnName,oz OriginalDatatype,oz Datatype,oz Column-index,lz-oz Datatype-Status,lz-oz Column-Status"
  # +" "+",sz SchenamName,sz TableName,sz ColumnName,sz OriginalDatatype,sz Datatype,sz Column-index,src-sz Datatype-Status,src-sz Column-Status"
  +"\n")


  count,srclz_failed,srclz_passed,lzoz_passed,lzoz_failed =0,0,0,0,0
  with f_lock :
    with f:
      reader = XLSXDictReader(f)
      templist = list(reader[1])

      from itertools import groupby
      srcservergrouped = groupby(templist,key=lambda row:(row['DatabaseType'],row['S_Server']))

      for key,rows in srcservergrouped :
        if key[0] =='MSSQL' :
          initialize_meta_mssql(key[1]) 
        
        for row in rows :
          # outputfile.write("id,SchenamName,TableName,ColumnName,OriginalDatatype,Datatype,Column-index," ",SchenamName,TableName,ColumnName,OriginalDatatype,Datatype,Column-index,Datatype-Status,Column-Status\n")
          # outputfile.write("\n")
          if row['ACTIVE'] == 'N':
            # print(row['S_TableName'] + "Skipped") 
            continue;
          else:

            if key[0] =='MSSQL' :
              # print(row['S_TableName'])
              mssql_columns_records = get_mssql_structure(row['S_Database'],row['S_Schema'],row['S_TableName'])
              src_lower_cols = transform_sqldatatype(lowercase_columns(mssql_columns_records))
            elif key[0] == 'Netezza':
              # print(row['S_TableName'])
              neteeza_columns_records = get_neteeza_structure(row['S_TableName'],neteeza_curs)
              src_lower_cols = lowercase_columns(neteeza_columns_records)



            lz_hive_columns_records = get_hive_structure_v2(row['LZ_SchemaName'],row['HiveTableName'],hive_curs) 
            lz_hive_lower_cols = lowercase_columns(lz_hive_columns_records)

            # if table missing just add table is missing

            oz_hive_columns_records = get_hive_structure_v2(row['Oz_SchemaName'],row['HiveTableName'],hive_curs)
            oz_hive_lower_cols = lowercase_columns(oz_hive_columns_records)


            # sz_hive_columns_records = get_hive_structure(row['Sz_SchemaName'],row['SZ_HiveTableName'])
            # sz_hive_lower_cols = lowercase_columns(sz_hive_columns_records)

            # print(do_comparison(sql_lower_cols,hive_lower_cols))
            import re
            src_lz_compare = do_comparison(src_lower_cols,lz_hive_lower_cols,row['Source'])
            lz_oz_compare = do_comparison(lz_hive_lower_cols,oz_hive_lower_cols,row['Source'])
            # oz_sz_compare = do_comparison(sql_lower_cols,sz_hive_lower_cols)
            # src_lz_outputfile.write(re.sub(r'^',row['Source'],src_lz_compare))
            # lz_oz_outputfile.write(re.sub(r'^',row['Source'],lz_oz_compare))

            src_lz_outputfile.write(src_lz_compare)
            lz_oz_outputfile.write(lz_oz_compare)

            count += 1

            if "MisMatch"  in src_lz_compare or "missing"  in src_lz_compare :
              srclz_failed += 1
            else :
              srclz_passed +=1
            if "MisMatch"  in lz_oz_compare or "missing"  in lz_oz_compare:
              lzoz_failed += 1
            else :
              lzoz_passed += 1
            # row['SZ_HiveTableName'] == '' or len(row['SZ_HiveTableName']) == 0 or row['SZ_HiveTableName']
            # row['Sz_SchemaName'],row['SZ_HiveTableName']


            
            print(row['S_TableName']+": Completed")

      src_lz_outputfile.write("\n================================ LZ Metadata Validation Execution Summary ===============================================\n")
      src_lz_outputfile.write(" Total number of Tables Validated                    : " + str(count)+"\n")
      src_lz_outputfile.write(" Number of Tables passed Metadata Validation : " + str(srclz_passed)+"\n")
      src_lz_outputfile.write(" Number of Tables Failed Metadata Validation : " + str(srclz_failed)+"\n")
      src_lz_outputfile.write("\n")
      src_lz_outputfile.write(" Input parameter file        : " + ParasFilePath+"\n")
      # src_lz_outputfile.write(" Raw file contains header?   : " + Header)
      src_lz_outputfile.write(" Output File has been stored : " + src_lz_MetaOutputFilename+"\n")
      src_lz_outputfile.write(" Environment selected        : " + Environment+"\n") #change to environment name
      src_lz_outputfile.write(" Execution started Time      : " + ExecuStartTime+"\n")
      src_lz_outputfile.write(" Execution End Time          : " + str(datetime.datetime.now())+"\n")
      src_lz_outputfile.write("\n================================================================================================\n")

      lz_oz_outputfile.write("\n================================ OZ Metadata Validation Execution Summary ===============================================\n")
      lz_oz_outputfile.write(" Total number of Tables Validated                    : " + str(count)+"\n")
      lz_oz_outputfile.write(" Number of Tables passed Metadata Validation : " + str(lzoz_passed)+"\n")
      lz_oz_outputfile.write(" Number of Tables Failed Metadata Validation : " + str(lzoz_failed)+"\n")
      lz_oz_outputfile.write("\n")
      lz_oz_outputfile.write(" Input parameter file        : " + ParasFilePath+"\n")
      # lz_oz_outputfile.write(" Raw file contains header?   : " + Header)
      lz_oz_outputfile.write(" Output File has been stored : " + lz_oz_MetaOutputFilename+"\n")
      lz_oz_outputfile.write(" Environment selected        : " + Environment+"\n") #change to environment name
      lz_oz_outputfile.write(" Execution started Time      : " + ExecuStartTime+"\n")
      lz_oz_outputfile.write(" Execution End Time          : " + str(datetime.datetime.now())+"\n")
      lz_oz_outputfile.write("\n================================================================================================\n")

      print("\n================================ Execution Summary ===============================================\n")
      print(" Total number of Tables Validated                    : " + str(count-1))
      print(" Number of Tables passed SRC LZ Metadata Validation : " + str(srclz_passed))
      print(" Number of Tables Failed SRC LZ Metadata Validation : " + str(srclz_failed))
      print(" Number of Tables passed LZ OZ  Metadata Validation : " + str(lzoz_passed))
      print(" Number of Tables Failed LZ OZ  Metadata Validation : " + str(lzoz_failed))
      print("\n")
      print(" Input parameter file        : " + ParasFilePath)
      # print(" Raw file contains header?   : " + Header)
      print(" Output File has been stored : " + src_lz_MetaOutputFilename+" "+lz_oz_MetaOutputFilename )
      print(" Environment selected        : " + Environment)
      print(" Execution started Time      : " + ExecuStartTime)
      print(" Execution End Time          : " + str(datetime.datetime.now()))
      print("\n================================================================================================\n")

      # close_hivemeta()
      close_hive()
      close_meta_mssql()
      src_lz_outputfile.close()
      lz_oz_outputfile.close()


if __name__ == "__main__" :  
  while True:
    pgm_choice = raw_input("=================== Lightning - Test Automation Framework ====================\n    SRC - LZ STAT Validation (1)\n    LZ and OZ Hist Count Validation (2)\n    LZ and OZ Metadata Validation (3)\n    LZ and OZ CDC Time Update (4)\n    LZ and OZ Incremental Count Validation (5)\n    LZ-OZ STAT Validation (6) \n    LZ-OZ Log File Validation (7) \n    SZ Count Validation (8)\n    SZ Metadata Validation (9)\n    SQL DW Count Validation (10)\n    SQL DW Metadata Validation (11)\n    SZ-SQL DW STAT Validation (12)\n    Denodo Count Validation (13)\n    End to End Regression Historical(14)\n    End to End Regression Incremental(15)\n\nPlease enter your choice :\n")
    if pgm_choice == "" or int(pgm_choice) >15 or int(pgm_choice) < 1:
      print " Please provide valid choice !!!"
      continue
    else:
      break
  if int(pgm_choice) == 1:
    STATMain(1) 
  elif int(pgm_choice) == 2 :
    do_CountValidation("hist")
  elif int(pgm_choice) == 3 :
    do_MetadataValidation()
  elif int(pgm_choice) == 4 :
    do_CDCUpdate()
  elif int(pgm_choice) == 5 :
    do_CountValidation("incremental")
  elif int(pgm_choice) == 6 :
    STATMain(2)
  elif int(pgm_choice) == 7 :    
    log_validation("Landing")
  elif int(pgm_choice) == 8 :    
    sz_count_validation()
  elif int(pgm_choice) == 9 :    
    do_SZMetadataValidation()
  elif int(pgm_choice) == 10 :    
    sz_azure_count_validation()
  elif int(pgm_choice) == 11 :    
    sz_azure_metadata_validation()
  elif int(pgm_choice) == 12 :
    STATMain(3)
  elif int(pgm_choice) == 13 :    
    do_denodocount()
  elif int(pgm_choice) == 14 :    
    do_AllValidationHist()
  elif int(pgm_choice) == 15 :    
    do_AllValidationIncr()




# =================== Lightning - Test Automation Framework ====================
#     LZ and OZ CDC Time Update (1)
#     LZ and OZ Hist Count Validation (2)
#     LZ and OZ Incremental Count Validation (3) 
#     LZ and OZ Metadata Validation (4)
#     SZ Count Validation (5)
#     SZ Metadata Validation (6)
#     SQl DW Count Validation (7)
#     SQl DW Metadata Validation (8)
#     Denodo Count Validation (9)
#     End to End Regression (10)

# Please enter your choice :
