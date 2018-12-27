import jaydebeapi
import jpype

import os
import sys
import csv
import datetime
import pymssql
from string import Template
import pprint
import re
import openpyxl
import pyodbc
from pprint import pprint
from dateutil.parser import parse
# Numeric Types
# =============
# tinyint
# smallint
# int
# integer
# bigint
# float
# double
# double precision
# decimal
# numeric

# -- min,max(<columnname>) as max_<columnname>,sum(<columnname>) as sum_<columnname>,avg(<columnname>) as avg_<columnname>

# Date/Time Types
# ===============
# timestamp
# date
# interval

# -- min,max(<columnname>) as max_<columnname>,avg(<columnname>) as avg_<columnname>

# String Types
# ============
# string
# varchar
# char

# -- min,max(<columnname>) as max_<columnname>,avg(<columnname>) as avg_<columnname> string length

# Misc Types
# ==========
# boolean
# -- count of false ,true
# binary


# Complex Types
# =============
# arrays
# maps
# structs
# union
def XLSXDictReader(f):
    book = openpyxl.reader.excel.load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))
    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)
    return (headers.values(),(dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1)))

def printsummary(validation,count,passed,failed,ParasFilePath,countOutputFilename,ExecuStartTime) :
  ostring ="\n================================ "+validation+" Validation Summary ===============================================\n"\
   +" \nTotal number of Tables Validated                    : " + str(count)\
   +" \nNumber of Tables that passed "+validation+" validation : " + str(passed)\
   +" \nNumber of Tables that failed "+validation+" validation : " + str(failed)\
   +"\n"\
   +" \nInput parameter file        : " + ParasFilePath\
   +" \nOutput File has been stored : " + countOutputFilename\
   +" \nEnvironment selected        : " + Environment\
   +" \nExecution started Time      : " + ExecuStartTime\
   +" \nExecution End Time          : " + str(datetime.datetime.now())\
   +"\n================================================================================================\n"
  return ostring
Environment = os.environ['env'] if os.environ['env'] else os.environ['ENV']
def STATMain(choice):
  print "\n=========================== STAT Validation =====================================\n";
  while True:
    InputFileName = raw_input(" Enter the input parameter file path with filename : ")
    ParasFilePath = InputFileName
    if ParasFilePath == "":
      print " Please provide the Input parameter file !!!"
      continue
    else:
      break
  with open(ParasFilePath, 'r') as f:
    if choice == 1:
      SRCLZStatVal(f,ParasFilePath)
    elif choice == 2:
      LZOZStatVal(f,ParasFilePath)
    else :
      SZAzureStatVal(f,ParasFilePath)

ms_env = {
'xle-edw-build-dev' : { 'host' : '10.66.24.181:1436','username' : 'PMXLEDWDev', 'pass' : 'PMXLEDWDev' } ,\
'xle-mdm-hub-test' : { 'host' : '10.67.24.72:1436','username' : 'HDXLDEEPDev', 'pass' : 'SEcu$#re12#$%' } ,\
'xli-idw-build-qa' : { 'host' : '10.65.24.201:1436','username' : 'sqxledeep', 'pass' : '2sD9JxQM' } ,\
'xl-emiedb-analytics-dev' : { 'host' : '10.66.24.52:1436' ,'username' : 'PMXLDEEPDev', 'pass' : 'YcD4fFa5' } ,\
'xli-idw-stage-dev' : { 'host' : '10.66.24.45:1436','username' : 'PMXliIDWDev', 'pass' : 'PMXliIDWDev' } ,\
'xle-edw-stage-dev' : { 'host' : '10.66.24.155:1436','username' : 'PMXLEDWDev', 'pass' : 'PMXLEDWDev' } ,\
'EWRDB0408D' : { 'host' : '10.66.24.111:1436','username' : 'sqdeeptest', 'pass' : '!@#$sqdeeptest!@#$' },\
'xle-edw-stage-dev' : { 'host' : '10.66.24.155:1436','username' : 'PMXLEDWDev', 'pass' : 'PMXLEDWDev' } ,\
'xle-mdm-hub-qa' : { 'host' : '10.65.24.171:1436','username' : 'sqdeeptest', 'pass' : '$qdee4test' },\
'preprod' : { 'host' : 'xlc-azu-eus2-ppd-edsppd-dw-server.database.windows.net','username' : 'dwadmin', 'pass' : '22Ride@uber', 'database' : 'dw_xle_sz' },

}


def Initialize_AzureDW():
  global AzureDW_cnxn,AzureDW_cur
  server = 'xlc-azu-eus2-eds-dev-dw-server.database.windows.net'
  database = 'dw_xle_sz'
  username = 'dwadmin'
  password = 'NZ2h$EN#'
  driver= '{ODBC Driver 17 for SQL Server}'
  AzureDW_cnxn = pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)    
  AzureDW_cur = AzureDW_cnxn.cursor()
  # print('Connected to Azure_DW')


def close_AzureDW():
  global AzureDW_cnxn,AzureDW_cur
  AzureDW_cur.close()
  AzureDW_cnxn.close()  

def initialize_meta_mssql(servername):
  global ms_conn,ms_cursor
  ms_conn = pymssql.connect(server=ms_env[servername]['host'],user=ms_env[servername]['username'], password=ms_env[servername]['pass'])
  ms_cursor = ms_conn.cursor()
  # print("Connected to MSSQL")

def close_meta_mssql():
      global ms_conn,ms_cursor
      ms_cursor.close()
      ms_conn.close()

hive_datatype_qry_translation = {
'tinyint' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'smallint' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'int' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'integer' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'bigint' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'float' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'double' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'double precision' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'decimal' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],
'numeric' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>','sum(cast(<columnname> as BIGINT)) as sum_<columnname>','avg(cast(<columnname> as BIGINT)) as avg_<columnname>'],

'timestamp' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>'],
'date' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>'],
'interval' : ['min(<columnname>) as min_<columnname>','max(<columnname>) as max_<columnname>'],

'string' : ['min(length(<columnname>)) as min_<columnname>','max(length(<columnname>)) as max_<columnname>','avg(length(<columnname>)) as avg_<columnname>'],
'varchar' : ['min(length(<columnname>)) as min_<columnname>','max(length(<columnname>)) as max_<columnname>','avg(length(<columnname>)) as avg_<columnname>'],
'char' : ['min(length(<columnname>)) as min_<columnname>','max(length(<columnname>)) as max_<columnname>','avg(length(<columnname>)) as avg_<columnname>'],

'boolean' : ['sum(case when bool_col then 1 end) as true_count','sum(case when not bool_col then 1 end) false_count','count(<columnname>)) as total_count']

}

# cast(regexp_replace(concat(coalesce(entdate,00000000),regexp_replace(enttime,' ',0),'0'),'(\\d{4})(\\d{2})(\\d{2})(\\d{2})(\\d{2})(\\d{2})(\\d{3})','$1-$2-$3 $4:$5:$6.$7') as timestamp) 
# cast('2018-06-05' as date)
sql_hive_conversion_lz = {
'varchar' : ('string', ['min(length(<columnname>)) as min_<columnname>','max(length(<columnname>)) as max_<columnname>','avg(length(<columnname>)) as avg_<columnname>']),
'nvarchar' : ('string', ['min(length(<columnname>)) as min_<columnname>','max(length(<columnname>)) as max_<columnname>','avg(length(<columnname>)) as avg_<columnname>']),
'datetime' : ('string', ['min(cast(<columnname> as date)) as min_<columnname>','max(cast(<columnname> as date)) as max_<columnname>','avg(cast(<columnname> as date)) as avg_<columnname>']),
'date' : 'string', 
'binary' : 'string',
'char' : 'string',
'datetime2' : 'string',
'decimal' : 'double',
'smallint' : 'int',
'bit' : 'boolean',
'float' :'double',
'tinyint' :'int',
'numeric':'string',
'time':'string',
'text':'string',
'timestamp' : 'string',
'money' : 'string'
}





hive_meta = {
'acenmetasqlserver01d' : { 'host' : 'acenmetasqlserver01d','username' : 'deeprd@acenmetasqlserver01d', 'pass' : 'XLC33readonly', 'database' : 'acenmetasqlserver01d' }
}

hive_string = Template("""SELECT 
DBS.NAME AS TABLE_SCHEMA,
TBLS.TBL_NAME AS TABLE_NAME,
COLUMNS_V2.INTEGER_IDX +1 AS INTEGER_IDX,
COLUMNS_V2.COLUMN_NAME AS COLUMN_NAME,
COLUMNS_V2.TYPE_NAME AS COLUMN_DATA_TYPE_DETAILS
FROM DBS
JOIN TBLS ON DBS.DB_ID = TBLS.DB_ID
JOIN SDS ON TBLS.SD_ID = SDS.SD_ID
JOIN COLUMNS_V2 ON COLUMNS_V2.CD_ID = SDS.CD_ID
WHERE TBLS.TBL_NAME='${tablename}'
AND DBS.NAME='${schemaname}'
ORDER BY INTEGER_IDX;""")

def initialize_hive():
  global hive_conn,hive_curs
  jHome = jpype.getDefaultJVMPath()
  class_path = str.join(":", ["../driver_jars/"+name for name in os.listdir("../driver_jars/")])
  jpype.startJVM(jHome, '-ea',  '-Djava.class.path='+class_path)

  #configuring log4j,otherwise will throw warning saying it isn't configured
  log4j = jpype.JPackage('org.apache.log4j')

  log4j.BasicConfigurator.configure()
  #setting the rootlogger to display only error messages
  log4j_logger = log4j.Logger.getRootLogger()
  log4j_logger.setLevel(log4j.Level.ERROR)

  driverName = "org.apache.hive.jdbc.HiveDriver";

  #project and environement specific variables

  # url = 'jdbc:hive2://10.186.224.51:10001/default;transportMode=http;httpPath=/hive2'

  # new server for hive https://acen11-deep-hdi-hive04-d.azurehdinsight.net/
  # url = 'jdbc:hive2://10.185.195.31:10001/default;transportMode=http;httpPath=/hive2'
  url = 'jdbc:hive2://10.185.135.56:10001/default;transportMode=http;httpPath=/hive2'
  # user = 'sshuser'
  # passd ='D33ppassword123!'
  user = 'infa'
  passd = 'aA1234567890'

  hive_conn = jaydebeapi.connect(driverName,url,[user, passd])
  hive_curs = hive_conn.cursor()

def close_hive() :
  global hive_conn,hive_curs
  hive_curs.close()
  hive_conn.close()

def convert_type(table):

  for k,v in table.items():
    # print("test")
    # print(type(v))
    if str == type(v) :
      y =""
      try:
        if '.' in v:
          table[k] = datetime.datetime.strptime(v,'%Y-%m-%d %H:%M:%S.%f')
          
        else :
          y = v + '.000000'
          # print(y)
          try :
            table[k] = datetime.datetime.strptime(y,'%Y-%m-%d %H:%M:%S.%f')
          except :
            v = v + ' 00:00:00.000000'
            # print(v)
            table[k] = datetime.datetime.strptime(v,'%Y-%m-%d %H:%M:%S.%f')
          # print(table[k],type(v))
      except Exception  as e :
        # print(parse(v))
        print(str(e))


  # print(table[k],type(table[k]))
  return table


def get_hive_structure(schemaname,tablename) :
    global hive_meta_conn,hive_meta_cursor
    hive_meta_cursor.execute(hive_string.substitute({'tablename':tablename ,'schemaname': schemaname}))
    return hive_meta_cursor

def initialize_hivemeta(servername=None):
    global hive_meta_conn,hive_meta_cursor
    if servername is None :
        servername = 'acenmetasqlserver01d' # the default sql db in azure for Hivemetastore

    # print(hive_meta[servername]['host']+","+hive_meta[servername]['username']+","+ hive_meta[servername]['pass']+","+hive_meta[servername]['database'])
    hive_meta_conn = pymssql.connect(server=hive_meta[servername]['host'],user=hive_meta[servername]['username'], password=hive_meta[servername]['pass'],database=hive_meta[servername]['database'])
    hive_meta_cursor = hive_meta_conn.cursor(as_dict=True)

def close_hivemeta():
      global hive_meta_conn,hive_meta_cursor
      hive_meta_conn.close()
      hive_meta_cursor.close() 

def getRowsAsDict(cursor) :
      import itertools
      desc = cursor.description
      column_names = [col[0] for col in desc]
      data = [dict(itertools.izip(column_names, row)) for row in cursor.fetchall()] 
      return data

def get_hive_structure_v2(schemaname,tablename,hive_curs) :
    # print("desc {}.{}".format(schemaname,tablename))
    try:
      hive_curs.execute("desc {}.{}".format(schemaname,tablename))    
      resultset = getRowsAsDict(hive_curs)
    except Exception as e:
      print(str(e))
      return dict() 

    #removing the partition section from the describe command
    try:
        target_index = resultset.index({u'col_name': u'', u'comment': None, u'data_type': None})
    except ValueError, e:
        target_index = None
    # {u'col_name': u'dw_batch_id', u'comment': u'', u'data_type': u'int'},

    resultset = resultset[:target_index]
    cnt = 1
    trans_result = []    
    for dicts in resultset :
        temp_dict = dict()
        temp_dict['TABLE_SCHEMA'] =schemaname
        temp_dict['TABLE_NAME'] =tablename
        temp_dict['COLUMN_NAME'] =dicts['col_name']
        temp_dict['COLUMN_DATA_TYPE_DETAILS'] =dicts['data_type']
        temp_dict['INTEGER_IDX'] =cnt
        trans_result.append(temp_dict)
        cnt += 1        
    return trans_result

def lowercase_columns(columns_records) :
    return_value = { row['COLUMN_NAME'].lower() : {k.lower(): str(row[k]).lower() for k in row } for row in columns_records  }
    return return_value
def do_comparison_stat(table_1,table_2,table_name):
  failflag = 0
  extra_keys_in_table1 = set(table_1.keys()) - set(table_2.keys())
  extra_keys_in_table2 = set(table_2.keys()) - set(table_1.keys())    
  output_string= ""
  if not table_1.keys() or not table_2.keys() :
    return ''
  cnt = 1
  # for k,v in table_1.items():
  #   print(k,v)
  c1,c2 = "",""
  thresh = datetime.timedelta(minutes = 1)
  for (k,v), (k2,v2) in zip(table_1.items(), table_2.items()):
    if str(k) == str(k2) :
      c1 = "Match"
    else :
      c1 = "Mismatch"
      failflag = 1
    if v == v2:
      c2 = "Match"
    else :
      # print(type(v),type(v2))
      if (type(v) is datetime.datetime and type(v2) is datetime.datetime) :
        if v2 > v:

          diff = v2 - v
        else :
          diff = v - v2
        if diff > thresh:
          c2 = "Mismatch"
          failflag = 1
        else :
          c2 = "Match"
      else :
        c2 = "Mismatch"
        failflag = 1
    output_string += "\""+str(cnt) \
        +"\",\""+table_name\
        +"\",\""+str(k)\
        +"\",\""+str(v)\
        +"\",\" "\
        +"\",\""+str(k2)\
        +"\",\""+str(v2)\
        +"\",\""+c1\
        + "\",\""+c2\
        +"\"\n"
        
    cnt += 1
  return (output_string,failflag)






def SZAzureStatVal(f,ParasFilePath) :

    initialize_hive()
    Initialize_AzureDW()
    reader = XLSXDictReader(f)
    templist = list(reader[1])
    print "\n============================== SZ AZURE STAT Validation In-progress ============================================\n";
    print ' SZ AZURE STAT Validation started at ' + str(datetime.datetime.now())
    ExecuStartTime = str(datetime.datetime.now())
    MetaOutputFilename = "../output/SZ_DW_SQL_STAT_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
    outputfile = open(MetaOutputFilename, 'a')
    outputfile.write("id,Table_Name,SZColumn,SZValue,,SQLDWColumn,SQLDWValue,ColumnComparison,ValueComparison"+"\n")
    count,passed,failed = 0,0,0
    comp = ""
    for rows in templist :
        if(rows['ACTIVE']=='N'):
          continue;
        else:

          sz_hive_columns_records = get_hive_structure_v2(rows['SZ_SchemaName'],rows['HiveTableName'],hive_curs)
          temp_string = ""
          hive_sql = ""
          azure_sql = ""
          # rows_returned = sz_hive_columns_records.fetchall()
          for idx,row in enumerate(sz_hive_columns_records) :
              # pprint.pprint(row)
              # print(idx)
              # print(len(rows_returned))
              # print(row['COLUMN_NAME'])
              # print(row['COLUMN_DATA_TYPE_DETAILS'])
              if '(' in row['COLUMN_DATA_TYPE_DETAILS'].lower() :
                  row['COLUMN_DATA_TYPE_DETAILS'] = re.sub(r'\(.*?\)','',row['COLUMN_DATA_TYPE_DETAILS'])
              for translations in hive_datatype_qry_translation[row['COLUMN_DATA_TYPE_DETAILS'].lower()] :
                  # translations.replace("<columnname>",row['COLUMN_NAME'])
                  temp_string +=translations.replace('<columnname>',row['COLUMN_NAME'],2)+","
              #remove the last comma from the sql
          hive_sql += "select "+re.sub(r'\,$','',temp_string)+" from "+rows['SZ_SchemaName']+"."+rows['HiveTableName']
          azure_sql += "select "+re.sub(r'\,$','',temp_string)+" from dbo"+"."+rows['HiveTableName']
          # print(azure_sql)
          try:
          #repalce length funtion to len
            azure_sql = azure_sql.replace("length","len")
            
            #print(hive_sql)
            hive_curs.execute(hive_sql)
            hive_cols = getRowsAsDict(hive_curs)
            AzureDW_cur.execute(azure_sql)
            azure_cols = getRowsAsDict(AzureDW_cur)
            dict_1,dict_2 = {},{}
            for x in hive_cols:
              dict_1 = x
              for y in azure_cols:
                dict_2 = y
                dict_2 = convert_type(dict_2)
                dict_1 = convert_type(dict_1)
                hive_azure_comp = do_comparison_stat(dict_2,dict_1,rows['HiveTableName'])
                azure_cols.remove(y)
                break
              if hive_azure_comp[1] == 0:
                comp = "Match"
                passed += 1
              else :
                comp = "Mismatch"
                failed += 1
              print(rows['HiveTableName'] +" : " + comp)
              # print(hive_azure_comp[0])
              # print(hive_azure_comp[0])
              
              outputfile.write(hive_azure_comp[0])
              outputfile.write("\n")
            count+=1
          except Exception as e:
            print(str(e))
    validation = "SZ_AZURE_STAT"
    outstr = ""
    outstr = printsummary(validation,count,passed,failed,ParasFilePath,MetaOutputFilename,ExecuStartTime) 
    outputfile.write(outstr)
    print(outstr)
    close_hive()
    close_AzureDW()



def SRCLZStatVal(f,ParasFilePath) :
    # initialize_meta_mssql('xle-edw-stage-dev')
    # initialize_hivemeta('acenmetasqlserver01d')
    reader = XLSXDictReader(f)
    templist = list(reader[1])
    initialize_hive()
    print "\n============================== SRC LZ STAT Validation In-progress ============================================\n";
    print ' SRC LZ STAT Validation started at ' + str(datetime.datetime.now())
    ExecuStartTime = str(datetime.datetime.now())
    
    # from itertools import groupby
    # srcservergrouped = groupby(templist,key=lambda row:(row['DB_Name'],row['Table_Name']))
    # lz_db   oz_db   tablename
    MetaOutputFilename = "../output/SRC_LZ_STAT_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"

    outputfile = open(MetaOutputFilename, 'a')
    outputfile.write("id,Table_Name,Src_Column,Src_ColValue,,LZColumn,LZColValue,ColumnComparison,ValueComparison"+"\n")
    count,passed,failed = 0,0,0
    comp = ""
    for rows in templist :

      initialize_meta_mssql(rows['S_Server'])
      lz_hive_columns_records = get_hive_structure_v2(rows['LZ_SchemaName'],rows['HiveTableName'],hive_curs)
      temp_string = ""
      hive_sql = ""
      SQL_sql = ""
      # rows_returned = sz_hive_columns_records.fetchall()
      for idx,row in enumerate(lz_hive_columns_records) :        
        # pprint.pprint(row)
        if row['COLUMN_NAME'] == 'loaddate' :
          continue
        # print(idx)
        # print(len(rows_returned))
        # print(row['COLUMN_NAME'])
        # print(row['COLUMN_DATA_TYPE_DETAILS'])
        if '(' in row['COLUMN_DATA_TYPE_DETAILS'].lower() :
            # print(row['COLUMN_DATA_TYPE_DETAILS'].lower())
            row['COLUMN_DATA_TYPE_DETAILS'] = re.sub(r'\(.*?\)','',row['COLUMN_DATA_TYPE_DETAILS'])
        for translations in hive_datatype_qry_translation[row['COLUMN_DATA_TYPE_DETAILS'].lower()] :
            # print(translations)
            # translations.replace("<columnname>",row['COLUMN_NAME'])
            temp_string +=translations.replace('<columnname>',row['COLUMN_NAME'],2)+","
        #remove the last comma from the sql
      hive_sql += "select "+re.sub(r'\,$','',temp_string)+" from "+rows['LZ_SchemaName']+"."+rows['HiveTableName']
      SQL_sql += "select "+re.sub(r'\,$','',temp_string)+" from "+ rows['S_Database']+"."+rows['S_Schema']+"."+rows['HiveTableName']
        # print(azure_sql)

        #repalce length funtion to len
      SQL_sql = SQL_sql.replace("length","len")
      # print(SQL_sql)
      # print(rows['HiveTableName'])
      print(hive_sql)
      try:
        hive_curs.execute(hive_sql)
      except Exception as e:
        print(str(e))
        continue
      

      # for col_val in getRowsAsDict(hive_curs)  :
      #   pprint(col_val)
      hive_cols = getRowsAsDict(hive_curs)
      ms_cursor.execute(SQL_sql)
      sql_cols = getRowsAsDict(ms_cursor)
      dict_1,dict_2 = {},{}

      for x in hive_cols:
        dict_1 = x
        for y in sql_cols:
          dict_2 = y
          dict_2 = convert_type(dict_2)
          dict_1 = convert_type(dict_1)
          SQL_LZ_COMP = do_comparison_stat(dict_2,dict_1,rows['HiveTableName'])
          sql_cols.remove(y)
          break
        if SQL_LZ_COMP[1] == 0:
          comp = "Match"
          passed += 1
        else :
          comp = "Mismatch"
          failed += 1

        print(rows['HiveTableName'] +" : " + comp)
        # print(hive_azure_comp[0])
        # print(SQL_LZ_COMP[0])
        
        outputfile.write(SQL_LZ_COMP[0])
        outputfile.write("\n")
      count+=1
    validation = "SQL_LZ_STAT"
    outstr = ""
    outstr = printsummary(validation,count,passed,failed,ParasFilePath,MetaOutputFilename,ExecuStartTime) 
    outputfile.write(outstr)
    print(outstr)
    # 
    # # MetaOutputFilePath+row['S_TableName']+
    # outputfile = open(MetaOutputFilename, 'w')

    # close_hivemeta()
    # close_meta_mssql()
    close_hive()
    close_meta_mssql()

def LZOZStatVal(f,ParasFilePath) :

    initialize_hive()
    reader = XLSXDictReader(f)
    templist = list(reader[1])
    MetaOutputFilename = "../output/LZ_OZ_STAT_"+datetime.datetime.now().strftime('%d-%b-%Y_%I:%M:%S_%p')+".csv"
    print "\n============================== LZ OZ STAT Validation In-progress ============================================\n";
    print ' LZ OZ STAT Validation started at ' + str(datetime.datetime.now())
    ExecuStartTime = str(datetime.datetime.now())
    outputfile = open(MetaOutputFilename, 'a')
    outputfile.write("id,Table_Name,LZ_Column,LZ_ColValue,,OZColumn,OZColValue,ColumnComparison,ValueComparison"+"\n")
    count,passed,failed = 0,0,0
    comp =""
    for rows in templist :
        oz_hive_columns_records = get_hive_structure_v2(rows['Oz_SchemaName'],rows['HiveTableName'],hive_curs)
        temp_string = ""
        lz_sql = ""
        oz_sql = ""
        # rows_returned = sz_hive_columns_records.fetchall()
        for idx,row in enumerate(oz_hive_columns_records) :
            # pprint.pprint(row)
            # print(idx)
            # print(len(rows_returned))
            # print(row['COLUMN_NAME'])
            # print(row['COLUMN_DATA_TYPE_DETAILS'])
            if '(' in row['COLUMN_DATA_TYPE_DETAILS'].lower() :
                row['COLUMN_DATA_TYPE_DETAILS'] = re.sub(r'\(.*?\)','',row['COLUMN_DATA_TYPE_DETAILS'])
            for translations in hive_datatype_qry_translation[row['COLUMN_DATA_TYPE_DETAILS'].lower()] :
                # translations.replace("<columnname>",row['COLUMN_NAME'])
                temp_string +=translations.replace('<columnname>',row['COLUMN_NAME'],2)+","
            #remove the last comma from the sql
        lz_sql += "select "+re.sub(r'\,$','',temp_string)+" from "+rows['LZ_SchemaName']+"."+rows['HiveTableName']
        oz_sql += "select "+re.sub(r'\,$','',temp_string)+" from "+rows['Oz_SchemaName']+"."+rows['HiveTableName']
        # print(lz_sql)
        # print(oz_sql)
        hive_curs.execute(lz_sql)
        lz_cols = getRowsAsDict(hive_curs)
        hive_curs.execute(oz_sql)
        oz_cols = getRowsAsDict(hive_curs)
        dict_1,dict_2 = {},{}
        for x in lz_cols:
          dict_1 = x
          for y in oz_cols:
            dict_2 = y
            dict_2 = convert_type(dict_2)
            dict_1 = convert_type(dict_1)
            LZ_OZ_COMP = do_comparison_stat(dict_2,dict_1,rows['HiveTableName'])
            oz_cols.remove(y)
            break
          if LZ_OZ_COMP[1] == 0:
            comp = "Match"
            passed += 1
          else :
            comp = "Mismatch"
            failed += 1
          print(rows['HiveTableName'] +" : " + comp)
          # print(hive_azure_comp[0])
          # print(LZ_OZ_COMP[0])         
          outputfile.write(LZ_OZ_COMP[0])
          outputfile.write("\n")
        count+=1
    validation = "LZ_OZ_STAT"
    outstr = ""
    outstr = printsummary(validation,count,passed,failed,ParasFilePath,MetaOutputFilename,ExecuStartTime) 
    outputfile.write(outstr)
    print(outstr)
    close_hive()

if __name__ == "__main__":
  STATMain(3)









